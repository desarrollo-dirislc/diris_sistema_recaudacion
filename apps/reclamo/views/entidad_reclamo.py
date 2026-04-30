from collections import defaultdict
import datetime
import io
import os
import sys
import datetime
from venv import logger
from django.views import View
import openpyxl
from django.utils.dateparse import parse_date

from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.decorators import method_decorator
from django.views.decorators.http import require_POST, require_http_methods
from django.utils import timezone
from django.core.mail import send_mail
from django.contrib import messages

from apps.reclamo.forms.entidad_reclamo_general import EntidadReclamoForm_general
from apps.reclamo.forms.entidad_reclamo_monitoreo import EntidadReclamoForm_monitoreo
from apps.reclamo.forms.entidad_reclamo_monitoreo_general import EntidadReclamoForm_monitoreo_general
from apps.reclamo.forms.entidad_reclamo_programacion import EntidadReclamoForm_programacion
from apps.reclamo.forms.entidad_reclamo_secretaria import EntidadReclamoForm_secretaria
from apps.reclamo.models.consolidado_diario import ConsolidadoDiario
from apps.reclamo.models.medida_adoptada import MedidaAdoptada
from apps.reclamo.models.monitoreo_internet import Monitoreo_internet
from apps.reclamo.models.programacion import Programacion
from apps.reclamo.models.sin_reclamo import SinReclamo
from apps.util.anios import ANIOS
from apps.util.update_menu import update_menu
from apps.util.valid_user_access_views import valid_access_view, permission_and_entidad, valid_ipress_entidad_add, \
    valid_ipress_entidad_edit
from main import settings
from setup.forms.clasificador_ingreso import ClasificadoresIngresoForm
from setup.forms.cuentas_contabilidad import ContabilidadCuentaForm
from setup.forms.servicios import SetupServiciosForm
from setup.models.clasificadores_ingreso import Clasificadores_ingreso
from setup.models.cuentas_contabilidad import ContabilidadCuenta
from setup.models.establecimientos import Establecimientos
from setup.models.servicios import SetupServicios

if os.path.splitext(os.path.basename(sys.argv[0]))[0] == 'pydoc-script':
    pass

from django.contrib import messages
from django.urls import reverse_lazy, reverse
from django.views.generic import CreateView, TemplateView, UpdateView, DeleteView, ListView

from apps.reclamo.forms.entidad_reclamo import EntidadReclamoForm, ReporteTicketForm
from apps.reclamo.models.clasificacion_causa import ClasificacionCausa
from apps.reclamo.models.entidad_reclamo import EntidadReclamo, MESES
from django.template.loader import render_to_string
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.shortcuts import render
from django.db import connection
from openpyxl.styles import Border, PatternFill, Font, Alignment, Side
import locale
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import requests
from django.http import JsonResponse
from django.views.decorators.http import require_POST 

from django.views import View
from django.shortcuts import redirect
from django.contrib import messages
from django.utils.decorators import method_decorator


from apps.reclamo.models.programacion import Programacion
from apps.util.valid_user_access_views import valid_access_view, valid_ipress_entidad_add
from datetime import datetime, time



from django.utils.dateparse import parse_date
from django.contrib.auth import get_user_model
from django.db import connection
from django.http import HttpResponse
from django.template.loader import render_to_string

 


 
paso_uno = [ 'tipo_incidencia', 'celular' ,'detalle_solicitud' ,'anydesk', 'correo_usuario','entidad2','n_atenciones','dependencia_service','usuario_service','dependencia_service_nombre','dependencia_padre','dependencia_padre_nombre','cargo_service','piso']
paso_dos = ['tipo_documento_usuario', 'numero_documento_usuario', 'razon_social_usuario', 'nombres_usuario',
            'apellido_paterno_usuario', 'apellido_materno_usuario']
paso_tres = [  'numero_documento_presenta', 'razon_social_presenta',
             'nombres_presenta', 'apellido_paterno_presenta', 'apellido_materno_presenta',
             'correo_presenta', 'domicilio_presenta', 'celular_presenta']
paso_cuatro = ['medio_recepcion_reclamo', 'fecha_reclamo', 'detalle_reclamo']
paso_cinco = ['servicio_hecho_reclamo', 'competencia_reclamo', 'clasificacion_reclamo_1',
              'clasificacion_reclamo_2', 'clasificacion_reclamo_3',  
              'codigo_reclamo_primigenio', 'etapa_reclamo', 'estado_reclamo', 'tipo_administrado_traslado',
              'codigo_administrado_deriva']

paso_cinco_prueba = ['servicio_hecho_reclamo', 'competencia_reclamo', 'clasificacion_reclamo_1', 'prioridad_clasificador_1',
              'clasificacion_reclamo_2',  'prioridad_clasificador_2' ,'clasificacion_reclamo_3', 'prioridad_clasificador_3', 
              'codigo_reclamo_primigenio',   'estado_reclamo', 'tipo_administrado_traslado',
              'codigo_administrado_deriva' ]

paso_seis = ['resultado_reclamo', 'motivo_conclusion_anticipada', 'fecha_resultado_reclamo',
             'comunicacion_resultado_reclamo', 'fecha_notificacion']


paso_seis_prueba = ['tiene_medida','resultado_reclamo', 'motivo_conclusion_anticipada', 'fecha_resultado_reclamo'
                    ]
paso_siete_prueba = ['comunicacion_resultado_reclamo', 'fecha_notificacion']

paso_uno_secretaria = ['nombres' ,'cargo','tipo_incidencia', 'celular' ,'detalle_solicitud'  ,'sede','nombre','anydesk','piso' ,'ris','entidad2','correo_usuario','n_atenciones','dependencia_service','usuario_service','dependencia_service_nombre','dependencia_padre','dependencia_padre_nombre','cargo_service'
            ]

paso_uno_general = ['nombres' ,'cargo','tipo_incidencia',  'celular' ,'detalle_solicitud'  ,'sede','nombre','anydesk','piso' ,'correo_usuario','ris','entidad2' 
            ]

paso_uno_programacion = [ 'distrito','detalle_programacion','fecha_programada' ,  'chofer' ,'auto' ,'descripcion_general','dependencia_service','dependencia_service_nombre'

            ]

paso_uno_monitoreo = ['ris','entidad2','distrito','velocidad_inter' ,'anexos', 'fecha_hora_perdida' ,'fecha_hora_reestablecido' ,'tiempo_sin_serv', 'estado', 'observacion' , 'tipo_incidencia' , 'correo' , 'evidencia','evidencia_2','cod_ticket_gtd'

            ]

paso_uno_monitoreo_general = ['ris','entidad2','distrito','velocidad_inter' ,'anexos', 'fecha_hora_perdida' ,'fecha_hora_reestablecido' ,'tiempo_sin_serv', 'estado', 'observacion', 'tipo_incidencia'  , 'correo', 'evidencia'

            ]


mantenimiento_clasificador_ingreso = ['codigo','descripcion']


mantenimiento_servicios = ['descripcion_servicio','precio']


mantenimiento_cuentas_contabilidad= ['descripcion_contabilidad','codigo_contabilidad']


inputs_hidden = ['created_at', 'updated_at',
                 'created_ip', 'updated_ip', 'periodo']



class EntidadReclamoList(ListView):
    model = EntidadReclamo
    paginate_by = 30

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        if len(self.request.user.groups.filter(name="Administrador IPRESS")) > 0:
            entidad_id = self.request.session['entidad_id']
            queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(nombres_usuario__icontains=query) | Q(apellido_paterno_usuario__icontains=query) |
                Q(apellido_materno_usuario__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
        queryset = queryset.filter(
    estado_reclamo=0,
    usuario_service__iexact=self.request.user.usuario_service
)
        return queryset.order_by('-id')
    

class EntidadReclamoList_soporte(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_soporte.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_soporte, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_soporte, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes
 
        #if len(self.request.user.groups.filter(name="Soporte")) > 0:
         #   entidad_id = self.request.session['entidad_id']
          #  queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(usuario_service_nombre__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
 
        queryset = queryset.filter(
            estado_reclamo=0,
            tipo_incidencia__in=[1,2,3, 4, 5,7]
        ).exclude(tipo_incidencia__in=[  6])
 
        return queryset.order_by('-id')
    
class EntidadReclamoList_soporte_encurso(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_soporte_encurso.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_soporte_encurso, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_soporte_encurso, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)
    

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        #if len(self.request.user.groups.filter(name="Soporte")) > 0:
         #   entidad_id = self.request.session['entidad_id']
          #  queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(usuario_service_nombre__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
            
        queryset = queryset.filter(
            Q(estado_reclamo=1) | Q(estado_reclamo=2),
 
            id_user=self.request.user.id,
            tipo_incidencia__in=[1,2,3, 4, 5,7]
        ).exclude(
            tipo_incidencia__in=[6]
        )
     
 
        return queryset.order_by('-id')
    

class EntidadReclamoList_soporte_cerrados(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_soporte_cerrados.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_soporte_cerrados, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_soporte_cerrados, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        #if len(self.request.user.groups.filter(name="Soporte")) > 0:
         #   entidad_id = self.request.session['entidad_id']
          #  queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(usuario_service_nombre__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
 

        queryset = queryset.filter(
            Q(estado_reclamo=3),
            id_user=self.request.user.id,
            tipo_incidencia__in=[1,2,3, 4, 5,7]
        ).exclude(
            tipo_incidencia__in=[6]
        )
     
 
        return queryset.order_by('-id')
    

class EntidadReclamoList_sihce(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_sihce.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_sihce, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_sihce, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        #if len(self.request.user.groups.filter(name="Soporte")) > 0:
         #   entidad_id = self.request.session['entidad_id']
          #  queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
               Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(usuario_service_nombre__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
 
   
        queryset = queryset.filter(
            estado_reclamo=0,
            tipo_incidencia__in=[2, 6]
        ).exclude(tipo_incidencia__in=[1, 3, 4, 5])
 
        return queryset.order_by('-id')
    
class EntidadReclamoList_sihce_encurso(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_sihce_encurso.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_sihce_encurso, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_sihce_encurso, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)
    

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        #if len(self.request.user.groups.filter(name="Soporte")) > 0:
         #   entidad_id = self.request.session['entidad_id']
          #  queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(usuario_service_nombre__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
             
           
        queryset = queryset.filter(
            Q(estado_reclamo=1) | Q(estado_reclamo=2),
            id_user=self.request.user.id,
            tipo_incidencia__in=[2,6]
        ).exclude(
            tipo_incidencia__in=[1, 3, 4,5]
         )
    
 
        return queryset.order_by('-id')
    

class EntidadReclamoList_sihce_cerrados(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_sihce_cerrados.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_sihce_cerrados, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_sihce_cerrados, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        #if len(self.request.user.groups.filter(name="Soporte")) > 0:
         #   entidad_id = self.request.session['entidad_id']
          #  queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(usuario_service_nombre__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
 
   
        queryset = queryset.filter(
            Q(estado_reclamo=3),
            id_user=self.request.user.id,
            tipo_incidencia__in=[2,6]
        ).exclude(
            tipo_incidencia__in=[1, 3, 4,5]
         )
    
 
        return queryset.order_by('-id')
    

class EntidadReclamoUpdate_soporte_sihce(UpdateView):
    model = EntidadReclamo
    form_class = EntidadReclamoForm_secretaria
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_soporte_sihce.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_soporte_sihce, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate_soporte_sihce,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
 
        msg = "Ticket actualizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Ticket"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate_soporte_sihce, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)    



class EntidadReclamoList_redes(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_sgd.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_redes, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los tickets"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_redes, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        #if len(self.request.user.groups.filter(name="Soporte")) > 0:
         #   entidad_id = self.request.session['entidad_id']
          #  queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
               Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(usuario_service_nombre__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
        

        queryset = queryset.filter(
            estado_reclamo=0,
            tipo_incidencia__in=[1,2,3,4,5,7]
        ).exclude(tipo_incidencia__in=[6])

        return queryset.order_by('-id')
    
class EntidadReclamoList_redes_encurso(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_sgd_encurso.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_redes_encurso, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los tickets"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_redes_encurso, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)
    

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        #if len(self.request.user.groups.filter(name="Soporte")) > 0:
         #   entidad_id = self.request.session['entidad_id']
          #  queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(usuario_service_nombre__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
            
         

        queryset = queryset.filter(
            Q(estado_reclamo=1) | Q(estado_reclamo=2),
            id_user=self.request.user.id,
            tipo_incidencia__in=[1,2,3,4,5,7]
        ).exclude(
            tipo_incidencia__in=[6]
        )
    
 
        return queryset.order_by('-id')
    

class EntidadReclamoList_redes_cerrados(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_sgd_cerrados.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_redes_cerrados, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_redes_cerrados, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()
        

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        #if len(self.request.user.groups.filter(name="Soporte")) > 0:
         #   entidad_id = self.request.session['entidad_id']
          #  queryset = queryset.filter(entidad_id=entidad_id)
 

 

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(usuario_service_nombre__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
            
       

        queryset = queryset.filter(
            Q(estado_reclamo=3),
            id_user=self.request.user.id,
            tipo_incidencia__in=[1,2,3,4,5,7]
        ).exclude(
            tipo_incidencia__in=[6]
        )
    
 
        return queryset.order_by('-id')
    

class EntidadReclamoUpdate_soporte_sgd(UpdateView):
    model = EntidadReclamo
    form_class = EntidadReclamoForm_secretaria
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_soporte_sgd.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_soporte_sgd, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate_soporte_sgd,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
 
            super(EntidadReclamoUpdate_soporte_sgd, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
             paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)    


 




class EntidadReclamoList_encurso(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_encurso.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_encurso, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_encurso, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        if len(self.request.user.groups.filter(name="Administrador IPRESS")) > 0:
            entidad_id = self.request.session['entidad_id']
            queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(nombres_usuario__icontains=query) | Q(apellido_paterno_usuario__icontains=query) |
                Q(apellido_materno_usuario__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
            
        queryset = queryset.filter(
            Q(estado_reclamo=1) | Q(estado_reclamo=2) , usuario_service__iexact=self.request.user.usuario_service  # <-- sigue filtrando por usuario logueado
)
    
 
        return queryset.order_by('-id')


class EntidadReclamoList_cerrados(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_cerrados.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_cerrados, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_cerrados, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        if len(self.request.user.groups.filter(name="Administrador IPRESS")) > 0:
            entidad_id = self.request.session['entidad_id']
            queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(nombres_usuario__icontains=query) | Q(apellido_paterno_usuario__icontains=query) |
                Q(apellido_materno_usuario__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
            
        queryset = queryset.filter(
            Q(estado_reclamo=3)  , usuario_service__iexact=self.request.user.usuario_service # <-- sigue filtrando por usuario logueado
)
    
 
        return queryset.order_by('-id')



class EntidadReclamoList_secretaria(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_secretaria.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_secretaria, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_secretaria, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        #if len(self.request.user.groups.filter(name="Soporte")) > 0:
         #   entidad_id = self.request.session['entidad_id']
          #  queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(usuario_service_nombre__icontains=query)  )

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
        queryset = queryset.filter(estado_reclamo=0)

        return queryset.order_by('-id')
    


class EntidadReclamoList_general(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_general.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_general, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"

        # Aquí NO pones valores por defecto, para detectar si el filtro viene o no
        anio = self.request.GET.get('anio')
        mes = self.request.GET.get('mes')
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], 
                anio=anio if anio else datetime.datetime.now().year, 
                periodo=mes if mes else datetime.datetime.now().month)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        # Agrega una bandera para el template si se están usando filtros
        filtros_usados = bool(query or anio or mes)

        context = super().get_context_data(**kwargs)
        context.update({
            'title': title,
            'anios': ANIOS,
            'meses': MESES,
            'anio': int(anio) if anio else '',
            'mes': int(mes) if mes else '',
            'query': query,
            'sin_reclamo': sin_reclamo,
            'show_results': filtros_usados,
        })
        return context


    def get_queryset(self):
        queryset = super().get_queryset()

        anio = self.request.GET.get('anio')
        mes = self.request.GET.get('mes')
        query = self.request.GET.get('query', "").strip()

        # Guardar en sesión solo si vienen definidos
        if anio:
            self.request.session['reclamo_anio'] = int(anio)
        if mes:
            self.request.session['reclamo_mes'] = int(mes)

        # Si NO hay filtros, devolver queryset vacío para que no muestre nada
        if not (query or anio or mes):
            return queryset.none()

        if query:
            queryset = queryset.filter(
               # Q(nombres__icontains=query) |
        Q(codigo_ticket=query)    # búsqueda exacta aquí
        #Q(nombres_usuario__icontains=query) |
        #Q(apellido_paterno_usuario__icontains=query) |
        #Q(apellido_materno_usuario__icontains=query)
            )

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=int(anio),
                fecha_reclamo__month=int(mes)
            )
        elif anio:
            queryset = queryset.filter(fecha_reclamo__year=int(anio))
        elif mes:
            queryset = queryset.filter(fecha_reclamo__month=int(mes))

        queryset = queryset.filter(estado_reclamo__in=[0, 1, 2,3])

        return queryset.order_by('-id')
    


    


class EntidadReclamoList_secretaria_encurso(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_secretaria_encurso.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_secretaria_encurso, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_secretaria_encurso, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        #if len(self.request.user.groups.filter(name="Soporte")) > 0:
         #   entidad_id = self.request.session['entidad_id']
          #  queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(usuario_service_nombre__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
        queryset = queryset.filter(
            Q(estado_reclamo=1) | Q(estado_reclamo=2),
 )

        return queryset.order_by('-id')

class EntidadReclamoList_secretaria_cerrados(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_secretaria_cerrados.html'
    

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList_secretaria_cerrados, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList_secretaria_cerrados, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        #if len(self.request.user.groups.filter(name="Soporte")) > 0:
         #   entidad_id = self.request.session['entidad_id']
          #  queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(nombres__icontains=query) | Q(codigo_ticket__icontains=query) |
                Q(usuario_service_nombre__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)
            
        queryset = queryset.filter(
            Q(estado_reclamo=3),
 )

        return queryset.order_by('-id')








class EntidadReclamoList1(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_prueba1.html'

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList1, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList1, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        if len(self.request.user.groups.filter(name="Prueba")) > 0:
            entidad_id = self.request.session['entidad_id']
            queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(codigo_administrado__icontains=query) | Q(razon_social_usuario__icontains=query) |
                Q(nombres_usuario__icontains=query) | Q(apellido_paterno_usuario__icontains=query) |
                Q(apellido_materno_usuario__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)

        # Filtro adicional para estado igual a 0
        queryset = queryset.filter(etapa_reclamo=1)
        # queryset = queryset.filter(Q(estado_reclamo__isnull=True))

        return queryset.order_by('-id')


class EntidadReclamoList2(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_prueba2.html'

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList2, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList2, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        if len(self.request.user.groups.filter(name="Prueba")) > 0:
            entidad_id = self.request.session['entidad_id']
            queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(codigo_administrado__icontains=query) | Q(razon_social_usuario__icontains=query) |
                Q(nombres_usuario__icontains=query) | Q(apellido_paterno_usuario__icontains=query) |
                Q(apellido_materno_usuario__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)

          # Filtro adicional para estado igual a 0
        queryset = queryset.filter(Q(estado_reclamo=2) | Q(
            estado_reclamo=3)  | Q(estado_reclamo=1)).exclude(etapa_reclamo=1)
        return queryset.order_by('-id')


class EntidadReclamoList3(ListView):
    model = EntidadReclamo
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_prueba3.html'

    @ method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoList3, self).dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        title = "Todos los reclamos"
        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        try:
            sin_reclamo = SinReclamo.objects.get(
                entidad_id=self.request.session['entidad_id'], anio=anio, periodo=mes)
        except SinReclamo.DoesNotExist:
            sin_reclamo = None

        return dict(
            super(EntidadReclamoList3, self).get_context_data(**kwargs), title=title, anios=ANIOS, meses=MESES,
            anio=anio, mes=mes, query=query, sin_reclamo=sin_reclamo)

    def get_queryset(self):
        queryset = super().get_queryset()

        anio = int(self.request.GET.get('anio', datetime.datetime.now().year))
        mes = int(self.request.GET.get('mes', datetime.datetime.now().month))
        query = self.request.GET.get('query', "")

        self.request.session['reclamo_anio'] = anio
        self.request.session['reclamo_mes'] = mes

        if len(self.request.user.groups.filter(name="Prueba")) > 0:
            entidad_id = self.request.session['entidad_id']
            queryset = queryset.filter(entidad_id=entidad_id)

        if query:
            queryset = queryset.filter(
                Q(codigo_administrado__icontains=query) | Q(razon_social_usuario__icontains=query) |
                Q(nombres_usuario__icontains=query) | Q(apellido_paterno_usuario__icontains=query) |
                Q(apellido_materno_usuario__icontains=query))

        if anio and mes:
            queryset = queryset.filter(
                fecha_reclamo__year=anio, fecha_reclamo__month=mes)

        # Filtro adicional  
        queryset = queryset.filter(
            Q(estado_reclamo=6) | Q(estado_reclamo=4) | Q(estado_reclamo=5) | Q(resultado_reclamo=5)
                        ).exclude(resultado_reclamo=0)
        
        

        return queryset.order_by('-id')


class EntidadReclamoCreate(CreateView):
    form_class = EntidadReclamoForm
    model = EntidadReclamo

    @method_decorator(valid_access_view(valid_ipress_entidad_add, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoCreate, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('reclamo:entidad-reclamo-list') + "?anio=" + str(
            self.request.session['reclamo_anio']) + "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoCreate, self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Ticket creado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        # No guardes todavía
        self.object = form.save(commit=False)

        # Asignar usuario logeado
        self.object.id_usuario = self.request.user.id
        self.object.nombres = self.request.user.first_name
        self.object.apellidos = self.request.user.last_name
        self.object.cargo_service_nombre = self.request.user.cargo
        self.object.dependencia_service_nombre = self.request.user.dependencia
        self.object.correo_usuario = self.request.user.email
        self.object.usuario_service = self.request.user.usuario_service
        self.object.usuario_service_nombre = (
            f"{self.request.user.first_name} {self.request.user.last_name}"
        )

        # Asignar otros campos
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id
        self.object.estado_reclamo = 0

        # Ahora sí guardar
        self.object.save()

        # ---------------- ENVIO DE CORREO ----------------
        correo_usuario = self.object.correo_usuario

        if correo_usuario:
            try:
                asunto = "Confirmación de registro de TICKET"
                mensaje = f"""
Estimado(a),

Su solicitud fue registrada correctamente en nuestro sistema.

Código de Ticket: {self.object.codigo_ticket}

Saludos cordiales,

Atte.

Oficina de Gestión de Tecnologías de la Información - DIRIS Lima Centro
"""
                send_mail(
                    subject=asunto,
                    message=mensaje,
                    from_email=settings.EMAIL_HOST_USER,
                    recipient_list=[correo_usuario],
                    fail_silently=False,
                )
            except Exception as e:
                messages.warning(
                    self.request,
                    "El reclamo fue registrado, pero no se pudo enviar el correo de confirmación."
                )
        # -------------------------------------------------

        return super().form_valid(form)

    def get_initial(self):
        initial = super(EntidadReclamoCreate, self).get_initial()
        if self.request.method == 'GET':
            initial.update(
                {'periodo': '000000', 'tipo_institucion': 1, 'tipo_documento_usuario': 1,
                 'tipo_documento_presenta': 1, 'competencia_reclamo': 1,
                 'autorizacion_notificacion_correo': 1})
        return initial

    def get_context_data(self, **kwargs):
        title = "Nuevo Reclamo"
        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoCreate, self).get_context_data(**kwargs),
            title=title,
            paso_uno=paso_uno,
            paso_dos=paso_dos,
            paso_tres=paso_tres,
            paso_cuatro=paso_cuatro,
            paso_cinco=paso_cinco,
            paso_seis=paso_seis,
            inputs_hidden=inputs_hidden,
            clasificaciones=clasificaciones
        )



class EntidadReclamoCreate_secretaria(CreateView):
    form_class = EntidadReclamoForm_secretaria
    model = EntidadReclamo
    template_name = 'reclamo/entidadreclamo_form_secretaria.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_add, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoCreate_secretaria, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('reclamo:secretaria-list-nuevos') + "?anio=" + str(
            self.request.session['reclamo_anio']) + "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoCreate_secretaria,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id

        return kwargs

    # def get_success_url(self):
    #     return reverse_lazy('ciclo:edit',
    #                         kwargs={'pk': self.object.id})

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        
        self.object.estado_reclamo = 0
        

        msg = f"Ticket creado correctamente. Código: {self.object.codigo_ticket}"
        messages.add_message(self.request, messages.SUCCESS, msg)

        correo_usuario = form.cleaned_data.get('correo_usuario')

        if correo_usuario:
            try:
                asunto = "Confirmación de registro de TICKET"
                mensaje = f"""
    Estimado(a),

    Su solicitud fue registrada correctamente en nuestro sistema.

    Código de Ticket: {self.object.codigo_ticket}

 
    Saludos cordiales,

    Atte.
 
    Oficina de Gestión de Tecnologías de la Información - DIRIS Lima Centro
     """
                send_mail(
                    subject=asunto,
                    message=mensaje,
                    from_email=settings.EMAIL_HOST_USER,
                    recipient_list=[correo_usuario],
                    fail_silently=False,
                )
            except Exception as e:
                messages.warning(self.request, "El reclamo fue registrado, pero no se pudo enviar el correo de confirmación.")
                # Aquí podrías loggear el error: print(e) o logging.error(e)

        return super().form_valid(form)
    
    

    def get_initial(self):
        initial = super(EntidadReclamoCreate_secretaria, self).get_initial()
        if self.request.method == 'GET':
            initial.update(
                {'periodo': '000000', 'tipo_institucion': 1, 'tipo_documento_usuario': 1, 'tipo_documento_presenta': 1,
                 'competencia_reclamo': 1, 'autorizacion_notificacion_correo': 1})
        return initial

    def get_context_data(self, **kwargs):
        title = "Nuevo Reclamo"
        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoCreate_secretaria, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno, paso_uno_secretaria=paso_uno_secretaria,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)
    



class EntidadReclamoCreate_programacion(CreateView):
    form_class = EntidadReclamoForm_programacion
    model = Programacion
    template_name = 'reclamo/entidadreclamo_form_programacion.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_add, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoCreate_programacion, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('reclamo:programacion-list-nuevos')
    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoCreate_programacion, self).get_form_kwargs(*args, **kwargs)
        return kwargs

    def form_valid(self, form):
        msg = "Programación creado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        # No guardar aún
        self.object = form.save(commit=False)
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id
        self.object.estado_programacion = 0  # O el campo correcto en Programacion
        self.object.save()

        return super().form_valid(form)
    
    

    def get_initial(self):
        initial = super(EntidadReclamoCreate_programacion, self).get_initial()
        if self.request.method == 'GET':
            initial.update(
                {'periodo': '000000', 'tipo_institucion': 1, 'tipo_documento_usuario': 1, 'tipo_documento_presenta': 1,
                 'competencia_reclamo': 1, 'autorizacion_notificacion_correo': 1})
        return initial

    def get_context_data(self, **kwargs):
        title = "Nuevo Reclamo"
        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoCreate_programacion, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno, paso_uno_programacion=paso_uno_programacion,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)

    def get_initial(self):
        initial = super(EntidadReclamoCreate_programacion, self).get_initial()
        if self.request.method == 'GET':
            initial.update(
                {'periodo': '000000', 'tipo_institucion': 1, 'tipo_documento_usuario': 1, 'tipo_documento_presenta': 1,
                 'competencia_reclamo': 1, 'autorizacion_notificacion_correo': 1})
        return initial

    def get_context_data(self, **kwargs):
        title = "Nuevo Reclamo"
        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoCreate_programacion, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno, paso_uno_programacion=paso_uno_programacion,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)
    
  




@method_decorator(valid_access_view(valid_ipress_entidad_add, login_url='/validate'), name='dispatch')
class EntidadReclamoCreate_programacion_entidad(View):

    def post(self, request, *args, **kwargs):

        entidad = request.user.entidad

        # 🔍 Buscar establecimiento
        establecimiento = Establecimientos.objects.filter(id=entidad.id).first()

        # 🧠 Obtener descripción (evita error si no existe)
        descripcion = establecimiento.descripcion if establecimiento else ""

        Programacion.objects.create(
            entidad_id=entidad.id,
            dependencia_service=entidad.id,
            dependencia_service_nombre=descripcion,  # 👈 AQUÍ
            estado_programacion=0
        )

        messages.success(request, "Nueva venta creada")

        return redirect(reverse_lazy('reclamo:ventas-x-entidad'))

 



class EntidadReclamoCreate_monitoreo(CreateView):
    form_class = EntidadReclamoForm_monitoreo
    model = Monitoreo_internet
    template_name = 'reclamo/entidadreclamo_form_monitoreo.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_add, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoCreate_monitoreo, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('reclamo:monitoreo-list-nuevos')
    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoCreate_monitoreo, self).get_form_kwargs(*args, **kwargs)
        return kwargs

    def form_valid(self, form):
        msg = "Ticket creado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        # No guardar aún
        self.object = form.save(commit=False)
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id
        
        if self.object.tipo_incidencia == 1:
            self.object.estado = 0
        if self.object.tipo_incidencia in [2, 3]:
            self.object.estado = 2
        if self.object.fecha_hora_reestablecido:
            self.object.estado = 1
        

        self.object.save()

        return super().form_valid(form)
    
    

    def get_initial(self):
        initial = super(EntidadReclamoCreate_monitoreo, self).get_initial()
        if self.request.method == 'GET':
            initial.update(
                {'periodo': '000000', 'tipo_institucion': 1, 'tipo_documento_usuario': 1, 'tipo_documento_presenta': 1,
                 'competencia_reclamo': 1, 'autorizacion_notificacion_correo': 1})
        return initial

    def get_context_data(self, **kwargs):
        title = "Nuevo Reclamo"
        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoCreate_monitoreo, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno, paso_uno_monitoreo=paso_uno_monitoreo,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)
    



class EntidadReclamoList_monitoreo(ListView):
    model = Monitoreo_internet
    context_object_name = "monitoreo"
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_monitoreo.html'

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_queryset(self):
        return Monitoreo_internet.objects.all().order_by('-id')

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        context = super().get_context_data(**kwargs)
        context["title"] = "Monitoreo"

    
        return context
    
class EntidadReclamoCreate_monitoreo_general(CreateView):
    form_class = EntidadReclamoForm_monitoreo_general
    model = Monitoreo_internet
    template_name = 'reclamo/entidadreclamo_form_monitoreo_general.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_add, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoCreate_monitoreo_general, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('reclamo:monitoreo-list-general')
    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoCreate_monitoreo_general, self).get_form_kwargs(*args, **kwargs)
        return kwargs

    def form_valid(self, form):
        

        

        # No guardar aún
        self.object = form.save(commit=False)
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id

        # ✅ Condiciones de estado
        if self.object.tipo_incidencia == 1:
            self.object.estado = 0
        elif self.object.tipo_incidencia in [2, 3]:
            self.object.estado = 2

        # Guardar el objeto primero (ya genera cod_ticket_monitoreo si lo hace el modelo)
        self.object.save()

        msg = f"Ticket creado correctamente. Código: {self.object.cod_ticket_monitoreo}"
        messages.add_message(self.request, messages.SUCCESS, msg)

        # ✅ Enviar correo si el campo correo existe
        if self.object.correo:
            asunto = "Código de Ticket - Monitoreo de Internet"
            mensaje = (
                f"Estimado usuario,\n\n"
                f"Su ticket ha sido registrado correctamente.\n\n"
                f"Código de ticket: {self.object.cod_ticket_monitoreo}\n\n"
                 

                f"Saludos cordiales,\n"
                f"Atte.\n"
 
                f"Oficina de Gestión de Tecnologías de la Información - DIRIS LIMA CENTRO"
 
                

    
            )
            remitente = settings.DEFAULT_FROM_EMAIL  # debe estar configurado en settings.py
            destinatario = [self.object.correo]

            try:
                send_mail(asunto, mensaje, remitente, destinatario)
            except Exception as e:
                messages.warning(self.request, f"No se pudo enviar el correo: {e}")

        return super().form_valid(form)
    
    

    def get_initial(self):
        initial = super(EntidadReclamoCreate_monitoreo_general, self).get_initial()
        if self.request.method == 'GET':
            initial.update(
                {'periodo': '000000', 'tipo_institucion': 1, 'tipo_documento_usuario': 1, 'tipo_documento_presenta': 1,
                 'competencia_reclamo': 1, 'autorizacion_notificacion_correo': 1})
        return initial

    def get_context_data(self, **kwargs):
        title = "Nuevo Reclamo"
        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoCreate_monitoreo_general, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno, paso_uno_monitoreo=paso_uno_monitoreo,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)
    

    
class EntidadReclamoList_monitoreo_general(ListView):
    model = Monitoreo_internet
    context_object_name = "monitoreo"
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_monitoreo_general.html'

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
    
    def get_queryset(self):
        queryset = Monitoreo_internet.objects.all().order_by('-id')
        cod_ticket = self.request.GET.get('cod_ticket_monitoreo')

        # ✅ Filtrar si el usuario busca por código
        if cod_ticket:
            queryset = queryset.filter(cod_ticket_monitoreo__icontains=cod_ticket)
        return queryset

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        context = super().get_context_data(**kwargs)
        context["title"] = "Monitoreo"

        # ✅ Para mantener el valor del campo en la plantilla
        context["cod_ticket_monitoreo"] = self.request.GET.get('cod_ticket_monitoreo', '')

        return context
     

class EntidadReclamoUpdate_monitoreo(UpdateView):
    model = Monitoreo_internet
    form_class = EntidadReclamoForm_monitoreo
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_monitoreo.html'

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_monitoreo, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:monitoreo-list-nuevos', kwargs={})

    def form_valid(self, form):
        msg = "Monitoreo actualizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        # Lógica de estado
        if self.object.tipo_incidencia == 1:
            self.object.estado = 0
        if self.object.tipo_incidencia in [2, 3]:
            self.object.estado = 2
        if self.object.fecha_hora_reestablecido:
            self.object.estado = 1

        # Guardamos el registro actualizado
        self.object = form.save()

        # -------------------------------------
        # ENVÍO DE CORREO SI CUMPLE CONDICIÓN
        # -------------------------------------
        if self.object.correo and self.object.fecha_hora_reestablecido:
            mensaje = f"""
Estimado(a),

Le informamos que su TICKET {self.object.cod_ticket_monitoreo} ha sido atendido correctamente.

Hora y Fecha de inicio de la incidencia: {self.object.fecha_hora_perdida}
Hora y Fecha de Reestablecimiento del servicio: {self.object.fecha_hora_reestablecido}

Gracias por comunicarte con nosotros.

Saludos cordiales.
ATTE.
 
Oficina de Gestión de Tecnologías de la Información - DIRIS LIMA CENTRO
 """
            try:
                send_mail(
                    subject="Actualización de Ticket de Monitoreo",
                    message=mensaje,
                    from_email=None,  # usa DEFAULT_FROM_EMAIL
                    recipient_list=[self.object.correo],
                    fail_silently=False
                )
            except Exception as e:
                print("Error enviando correo:", e)

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"
        return dict(
            super(EntidadReclamoUpdate_monitoreo, self).get_context_data(**kwargs),
            title=title,
            paso_uno=paso_uno,
            paso_uno_monitoreo=paso_uno_monitoreo,
            paso_dos=paso_dos,
            paso_tres=paso_tres,
            paso_cuatro=paso_cuatro,
            paso_cinco=paso_cinco,
            paso_seis=paso_seis,
            inputs_hidden=inputs_hidden,
        )




class EntidadReclamoList_programacion(ListView):
    model = Programacion
    context_object_name = "programaciones"
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_programacion.html'

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        context = super().get_context_data(**kwargs)
        context["title"] = "Programaciones"

        for p in context["programaciones"]:
            p.personal = obtener_personal_programacion(p.id)

        return context

    def get_queryset(self):
        queryset = (
            super()
            .get_queryset()
            .filter(estado_programacion=0)
            .order_by('-id')
        )

        # 🔍 BUSCADOR
        busqueda = self.request.GET.get('q')

        if busqueda:
            queryset = queryset.filter(
                dependencia_service_nombre__icontains=busqueda
            )

        # 📅 FILTROS DE FECHA
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')

        fecha_inicio_parseada = parse_date(fecha_inicio) if fecha_inicio else None
        fecha_fin_parseada = parse_date(fecha_fin) if fecha_fin else None

        # 🟡 VALIDACIÓN
        if fecha_inicio_parseada and fecha_fin_parseada:
            if fecha_inicio_parseada > fecha_fin_parseada:
                messages.warning(
                    self.request,
                    "La fecha de inicio no puede ser mayor que la fecha de fin."
                )
                return queryset.none()

        # 🔹 FILTRAR POR FECHA
        if fecha_inicio_parseada:
            queryset = queryset.filter(
                created_at__gte=datetime.combine(fecha_inicio_parseada, time.min)
            )

        if fecha_fin_parseada:
            queryset = queryset.filter(
                created_at__lte=datetime.combine(fecha_fin_parseada, time.max)
            )

        return queryset

class EntidadReclamoList_ventas_x_entidad(ListView):
    model = Programacion
    context_object_name = "programaciones"
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_ventas_x_entidad.html'

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        context = super().get_context_data(**kwargs)
        context["title"] = "Programaciones"

        for p in context["programaciones"]:
            p.personal = obtener_personal_programacion(p.id)

        return context

    def get_queryset(self):
        user = self.request.user

        queryset = (
            super()
            .get_queryset()
            .filter(
                estado_programacion=0,
                dependencia_service=user.entidad_id
            )
            .order_by('-id')
        )

        # 🔥 FILTROS
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')

        fecha_inicio_parseada = parse_date(fecha_inicio) if fecha_inicio else None
        fecha_fin_parseada = parse_date(fecha_fin) if fecha_fin else None

        # 🚨 VALIDACIÓN
        if fecha_inicio_parseada and fecha_fin_parseada:
            if fecha_inicio_parseada > fecha_fin_parseada:
                messages.warning(
    self.request,
    "La fecha de inicio no puede ser mayor que la fecha de fin."
)
                return queryset.none()  # 👈 no devuelve nada

        # 🔹 FILTRAR
        if fecha_inicio_parseada:
            fecha_inicio_datetime = datetime.combine(fecha_inicio_parseada, time.min)
            queryset = queryset.filter(created_at__gte=fecha_inicio_datetime)

        if fecha_fin_parseada:
            fecha_fin_datetime = datetime.combine(fecha_fin_parseada, time.max)
            queryset = queryset.filter(created_at__lte=fecha_fin_datetime)

        return queryset



class EntidadReclamoConsolidadoDiarioVentasListView(ListView):
    template_name = 'reclamo/entidadreclamo_consolidado_diario_ventas.html'
    context_object_name = 'lista'
    paginate_by = 20

    def get_queryset(self):
        entidad_id = self.request.user.entidad_id

        with connection.cursor() as cursor:
            cursor.callproc(
                'sp_mostrar_consolidado_diario',
                [entidad_id]
            )

            columnas = [col[0] for col in cursor.description]

            resultados = [
                dict(zip(columnas, fila))
                for fila in cursor.fetchall()
            ]

        # más actual primero
        resultados.sort(
            key=lambda x: x['fecha'],
            reverse=True
        )

        return resultados

    def post(self, request, *args, **kwargs):
        fecha = request.POST.get('fecha')

        if not fecha:
            messages.error(
                request,
                'Seleccione una fecha.'
            )
            return redirect(request.path)

        entidad_id = request.user.entidad_id

        # 🔥 VALIDAR SI YA EXISTE CONSOLIDADO
        existe = ConsolidadoDiario.objects.filter(
            establecimiento_id=entidad_id,
            fecha=fecha
        ).exists()

        if existe:
            messages.warning(
                request,
                'Ya se realizó el consolidado para esta fecha.'
            )
            return redirect(request.path)

        # 🔥 SI NO EXISTE, EJECUTA PROCEDURE
        with connection.cursor() as cursor:
            cursor.callproc(
                'sp_generar_consolidado_diario',
                [entidad_id, fecha]
            )

        messages.success(
            request,
            'Consolidado generado correctamente.'
        )

        return redirect(request.path)
    


class EliminarConsolidadoDiarioView(View):

    def post(self, request, pk, *args, **kwargs):
        registro = get_object_or_404(
            ConsolidadoDiario,
            pk=pk
        )

        registro.delete()

        messages.success(
            request,
            'Registro eliminado correctamente.'
        )

        return redirect('reclamo:consolidado-diario')
    



class EntidadReclamoConsolidadoDiarioVentasListView_admin(ListView):
    template_name = 'reclamo/entidadreclamo_consolidado_diario_ventas_admin.html'
    context_object_name = 'lista'
    paginate_by = 20

    def get_queryset(self):
        q = self.request.GET.get('q', '').strip()
        fecha_inicio = self.request.GET.get('fecha_inicio', '').strip()
        fecha_fin = self.request.GET.get('fecha_fin', '').strip()

        # 🔥 VALIDACIÓN FECHAS
        if fecha_inicio and fecha_fin:
            if fecha_inicio > fecha_fin:
                messages.warning(
                    self.request,
                    "La fecha de inicio no puede ser mayor que la fecha de fin."
                )
                return []

        with connection.cursor() as cursor:
            cursor.callproc(
                'sp_mostrar_consolidado_diario_admin'
            )

            columnas = [col[0] for col in cursor.description]

            resultados = [
                dict(zip(columnas, fila))
                for fila in cursor.fetchall()
            ]

        # 🔍 Buscar establecimiento
        if q:
            resultados = [
                row for row in resultados
                if q.lower() in str(
                    row.get('descripcion', '')
                ).lower()
            ]

        # 📅 Fecha inicio
        if fecha_inicio:
            resultados = [
                row for row in resultados
                if str(row['fecha']) >= fecha_inicio
            ]

        # 📅 Fecha fin
        if fecha_fin:
            resultados = [
                row for row in resultados
                if str(row['fecha']) <= fecha_fin
            ]

        # Orden descendente
        resultados.sort(
            key=lambda x: x['fecha'],
            reverse=True
        )

        return resultados

    def post(self, request, *args, **kwargs):
        fecha = request.POST.get('fecha')

        if not fecha:
            messages.error(
                request,
                'Seleccione una fecha.'
            )
            return redirect(request.path)

        entidad_id = request.user.entidad_id

        existe = ConsolidadoDiario.objects.filter(
            establecimiento_id=entidad_id,
            fecha=fecha
        ).exists()

        if existe:
            messages.warning(
                request,
                'Ya se realizó el consolidado para esta fecha.'
            )
            return redirect(request.path)

        with connection.cursor() as cursor:
            cursor.callproc(
                'sp_generar_consolidado_diario',
                [entidad_id, fecha]
            )

        messages.success(
            request,
            'Consolidado generado correctamente.'
        )

        return redirect(request.path)


class EntidadReclamoList_reportes_ventas_admin(ListView):
    model = Programacion
    context_object_name = "programaciones"
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_reporte_ventas_administrador.html'

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        context = super().get_context_data(**kwargs)
        context["title"] = "Programaciones"

        for p in context["programaciones"]:
            p.personal = obtener_personal_programacion(p.id)

        return context

    def get_queryset(self):
        queryset = (
            super()
            .get_queryset()
            .filter(estado_programacion=0)
            .order_by('-id')
        )

        # 🔍 BUSCADOR
        busqueda = self.request.GET.get('q')

        if busqueda:
            queryset = queryset.filter(
                dependencia_service_nombre__icontains=busqueda
            )

        # 📅 FILTROS DE FECHA
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')

        fecha_inicio_parseada = parse_date(fecha_inicio) if fecha_inicio else None
        fecha_fin_parseada = parse_date(fecha_fin) if fecha_fin else None

        # 🟡 VALIDACIÓN
        if fecha_inicio_parseada and fecha_fin_parseada:
            if fecha_inicio_parseada > fecha_fin_parseada:
                messages.warning(
                    self.request,
                    "La fecha de inicio no puede ser mayor que la fecha de fin."
                )
                return queryset.none()

        # 🔹 FILTRAR POR FECHA
        if fecha_inicio_parseada:
            queryset = queryset.filter(
                created_at__gte=datetime.combine(fecha_inicio_parseada, time.min)
            )

        if fecha_fin_parseada:
            queryset = queryset.filter(
                created_at__lte=datetime.combine(fecha_fin_parseada, time.max)
            )

        return queryset
    



class EntidadReclamoList_reportes_ventas_entidad(ListView):
    model = Programacion
    context_object_name = "programaciones"
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_reporte_ventas_entidad.html'

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)
        context = super().get_context_data(**kwargs)
        context["title"] = "Programaciones"

        for p in context["programaciones"]:
            p.personal = obtener_personal_programacion(p.id)

        return context

    def get_queryset(self):
        queryset = (
            super()
            .get_queryset()
            .filter(estado_programacion=0)
            .order_by('-id')
        )

        # 🔍 BUSCADOR
        busqueda = self.request.GET.get('q')

        if busqueda:
            queryset = queryset.filter(
                dependencia_service_nombre__icontains=busqueda
            )

        # 📅 FILTROS DE FECHA
        fecha_inicio = self.request.GET.get('fecha_inicio')
        fecha_fin = self.request.GET.get('fecha_fin')

        fecha_inicio_parseada = parse_date(fecha_inicio) if fecha_inicio else None
        fecha_fin_parseada = parse_date(fecha_fin) if fecha_fin else None

        # 🟡 VALIDACIÓN
        if fecha_inicio_parseada and fecha_fin_parseada:
            if fecha_inicio_parseada > fecha_fin_parseada:
                messages.warning(
                    self.request,
                    "La fecha de inicio no puede ser mayor que la fecha de fin."
                )
                return queryset.none()

        # 🔹 FILTRAR POR FECHA
        if fecha_inicio_parseada:
            queryset = queryset.filter(
                created_at__gte=datetime.combine(fecha_inicio_parseada, time.min)
            )

        if fecha_fin_parseada:
            queryset = queryset.filter(
                created_at__lte=datetime.combine(fecha_fin_parseada, time.max)
            )

        return queryset









class EntidadReclamoList_programacion_atendidas(ListView):
    model = Programacion
    context_object_name = "programaciones"

    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_programacion_atendidas.html'

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)  # si quieres mantener el menú dinámico
        context = super().get_context_data(**kwargs)
        context["title"] = "Programaciones"
        return context

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .filter(estado_programacion=1)   # 👈 solo atendidas
            .order_by('-id')
        )
    




    
class EntidadReclamoList_programacion_soporte(ListView):
    model = Programacion
    context_object_name = "programaciones"
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_programacion_soporte.html'

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)  # mantiene el menú dinámico
        context = super().get_context_data(**kwargs)
        context["title"] = "Programaciones (Soporte)"

        # 🔹 agregar campo personal a cada Programacion con el SP
        for p in context["programaciones"]:
            p.personal = obtener_personal_programacion(p.id)

        return context

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .filter(estado_programacion=0)   # 👈 solo atendidas
            .order_by('-id')
        )
class EntidadReclamoList_programacion_soporte_atendidas(ListView):
    model = Programacion
    context_object_name = "programaciones"
    paginate_by = 30
    template_name = 'reclamo/entidadreclamo_list_programacion_soporte_atendidos.html'

    @method_decorator(valid_access_view(permission_and_entidad, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        update_menu(self.request)  # mantiene el menú dinámico
        context = super().get_context_data(**kwargs)
        context["title"] = "Programaciones Atendidas"

        # 🔹 agregar campo personal a cada Programacion usando el SP
        for p in context["programaciones"]:
            p.personal = obtener_personal_programacion(p.id)

        return context

    def get_queryset(self):
        return (
            super()
            .get_queryset()
            .filter(estado_programacion=1)   # 👈 solo atendidas
            .order_by('-id')
        )



class EntidadReclamoCreate_general(CreateView):
    form_class = EntidadReclamoForm_general
    model = EntidadReclamo
    template_name = 'reclamo/entidadreclamo_form_general.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_add, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoCreate_general, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('reclamo:general-list')  

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoCreate_general,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id

        return kwargs

    # def get_success_url(self):
    #     return reverse_lazy('ciclo:edit',
    #                         kwargs={'pk': self.object.id})

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.save()
        #entidad_id = self.request.session['entidad_id']
        #self.object.entidad_id = entidad_id
        self.object.estado_reclamo = 0
        

        msg = f"Ticket creado correctamente. Código: {self.object.codigo_ticket}"
        messages.add_message(self.request, messages.SUCCESS, msg)

        correo_usuario = form.cleaned_data.get('correo_usuario')

        if correo_usuario:
            try:
                asunto = "Confirmación de registro de TICKET"
                mensaje = f"""
    Estimado(a),

    Su solicitud fue registrada correctamente en nuestro sistema.

    Código de Ticket: {self.object.codigo_ticket}

    

    Saludos cordiales,

    Atte.
 
    Oficina de Gestión de Tecnologías de la Información - DIRIS LIMA CENTRO
     """
                send_mail(
                    subject=asunto,
                    message=mensaje,
                    from_email=settings.EMAIL_HOST_USER,
                    recipient_list=[correo_usuario],
                    fail_silently=False,
                )
            except Exception as e:
                messages.warning(self.request, "El reclamo fue registrado, pero no se pudo enviar el correo de confirmación.")
                # Aquí podrías loggear el error: print(e) o logging.error(e)

        return super().form_valid(form)
        
    

    def get_initial(self):
        initial = super(EntidadReclamoCreate_general, self).get_initial()
        if self.request.method == 'GET':
            initial.update(
                {'periodo': '000000', 'tipo_institucion': 1, 'tipo_documento_usuario': 1, 'tipo_documento_presenta': 1,
                 'competencia_reclamo': 1, 'autorizacion_notificacion_correo': 1})
        return initial

    def get_context_data(self, **kwargs):
        title = "Nuevo Reclamo"
        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoCreate_general, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno, paso_uno_general=paso_uno_general,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)


class EntidadReclamoCreate_prueba(CreateView):
    #success_url = reverse_lazy('reclamo:entidad-reclamo-registro/list')
    form_class = EntidadReclamoForm
    model = EntidadReclamo
    template_name = 'reclamo/entidadreclamo_form_prueba1.html'

 
    @method_decorator(valid_access_view(valid_ipress_entidad_add, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoCreate_prueba, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('reclamo:entidad-reclamo-registro/list') + "?anio=" + str(
            self.request.session['reclamo_anio']) + "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoCreate_prueba,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id

        return kwargs

    # def get_success_url(self):
    #     return reverse_lazy('ciclo:edit',
    #                         kwargs={'pk': self.object.id})

    def form_valid(self, form):

        msg = "Ticket creado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id
        self.object.estado_reclamo = 2
        self.object.etapa_reclamo = 1

        return super().form_valid(form)

    def get_initial(self):
        initial = super(EntidadReclamoCreate_prueba, self).get_initial()
        if self.request.method == 'GET':
            initial.update(
                {'periodo': '000000', 'tipo_institucion': 1, 'tipo_documento_usuario': 1, 'tipo_documento_presenta': 1,
                 'competencia_reclamo': 1, 'autorizacion_notificacion_correo': 1})
        return initial

    def get_context_data(self, **kwargs):
        title = "Nuevo Reclamo"
        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoCreate_prueba, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)


class EntidadReclamoUpdate(UpdateView):
    model = EntidadReclamo
    form_class = EntidadReclamoForm
    template_name = 'reclamo/entidadreclamo_form_ver_detalle.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)
    

class EntidadReclamoUpdate_programacion(UpdateView):
    model = Programacion
    form_class = EntidadReclamoForm_programacion
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_programacion.html'

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_programacion, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return (
            reverse('reclamo:programacion-list-nuevos', kwargs={})
             
        )

    def form_valid(self, form):
        msg = "Programación actualizada correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)
        self.object.estado_programacion = 0 
        self.object = form.save()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"
        return dict(
            super(EntidadReclamoUpdate_programacion, self).get_context_data(**kwargs),
            title=title,
            paso_uno=paso_uno,
            paso_uno_programacion=paso_uno_programacion,
            paso_dos=paso_dos,
            paso_tres=paso_tres,
            paso_cuatro=paso_cuatro,
            paso_cinco=paso_cinco,
            paso_seis=paso_seis,
            inputs_hidden=inputs_hidden,
         )
    


class EntidadReclamoUpdate_programacion_atendidas(UpdateView):
    model = Programacion
    form_class = EntidadReclamoForm_programacion
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_programacion_atendidas.html'

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_programacion_atendidas, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return (
            reverse('reclamo:programacion-list-nuevos', kwargs={})
            + "?anio=" + str(self.request.session['reclamo_anio'])
            + "&mes=" + str(self.request.session['reclamo_mes'])
        )

    def form_valid(self, form):
        msg = "Programación actualizada correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)
        self.object.estado_programacion = 0 
        self.object = form.save()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"
        return dict(
            super(EntidadReclamoUpdate_programacion_atendidas, self).get_context_data(**kwargs),
            title=title,
            paso_uno=paso_uno,
            paso_uno_programacion=paso_uno_programacion,
            paso_dos=paso_dos,
            paso_tres=paso_tres,
            paso_cuatro=paso_cuatro,
            paso_cinco=paso_cinco,
            paso_seis=paso_seis,
            inputs_hidden=inputs_hidden,
         )






class EntidadReclamoUpdate_programacion_soporte(UpdateView):
    model = Programacion
    form_class = EntidadReclamoForm_programacion
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_programacion_soporte.html'

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_programacion_soporte, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return (
            reverse('reclamo:programacion-list-nuevos', kwargs={})
            + "?anio=" + str(self.request.session['reclamo_anio'])
            + "&mes=" + str(self.request.session['reclamo_mes'])
        )

    def form_valid(self, form):
        msg = "Programación actualizada correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)
        self.object.estado_programacion = 0 
        self.object = form.save()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"
        return dict(
            super(EntidadReclamoUpdate_programacion_soporte, self).get_context_data(**kwargs),
            title=title,
            paso_uno=paso_uno,
            paso_uno_programacion=paso_uno_programacion,
            paso_dos=paso_dos,
            paso_tres=paso_tres,
            paso_cuatro=paso_cuatro,
            paso_cinco=paso_cinco,
            paso_seis=paso_seis,
            inputs_hidden=inputs_hidden,
         )



class EntidadReclamoUpdate_programacion_atendidas_soporte(UpdateView):
    model = Programacion
    form_class = EntidadReclamoForm_programacion
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_programacion_atendidas.html'

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_programacion_atendidas_soporte, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return (
            reverse('reclamo:programacion-list-nuevos', kwargs={})
            + "?anio=" + str(self.request.session['reclamo_anio'])
            + "&mes=" + str(self.request.session['reclamo_mes'])
        )

    def form_valid(self, form):
        msg = "Programación actualizada correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)
        self.object.estado_programacion = 0 
        self.object = form.save()
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"
        return dict(
            super(EntidadReclamoUpdate_programacion_atendidas_soporte, self).get_context_data(**kwargs),
            title=title,
            paso_uno=paso_uno,
            paso_uno_programacion=paso_uno_programacion,
            paso_dos=paso_dos,
            paso_tres=paso_tres,
            paso_cuatro=paso_cuatro,
            paso_cinco=paso_cinco,
            paso_seis=paso_seis,
            inputs_hidden=inputs_hidden,
         )







class EntidadReclamoUpdate_soporte(UpdateView):
    model = EntidadReclamo
    form_class = EntidadReclamoForm_secretaria
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_soporte.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_soporte, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate_soporte,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate_soporte, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)


class EntidadReclamoUpdate_soporte_encurso(UpdateView):
    model = EntidadReclamo
    form_class = EntidadReclamoForm_secretaria
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_soporte_encurso.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_soporte_encurso, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate_soporte_encurso,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate_soporte_encurso, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)
    

class EntidadReclamoUpdate_secretaria(UpdateView):
    model = EntidadReclamo
    form_class = EntidadReclamoForm_secretaria
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_secretaria.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_secretaria, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate_secretaria,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate_secretaria, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)


class EntidadReclamoUpdate_secretaria_encurso(UpdateView):
    model = EntidadReclamo
    form_class = EntidadReclamoForm_secretaria
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_secretaria_encurso.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_secretaria_encurso, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate_secretaria_encurso,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate_secretaria_encurso, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)

class EntidadReclamoUpdate_secretaria_cerrados(UpdateView):
    model = EntidadReclamo
    form_class = EntidadReclamoForm_secretaria
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_secretaria_cerrados.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_secretaria_cerrados, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate_secretaria_cerrados,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate_secretaria_cerrados, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)





class EntidadReclamoUpdate_general(UpdateView):
    model = EntidadReclamo
    form_class = EntidadReclamoForm_secretaria
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_general.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_general, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate_general,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate_general, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)

 

class EntidadReclamoUpdate1(UpdateView):
    success_url = reverse_lazy('reclamo:entidad-reclamo-registro/list')
    model = EntidadReclamo
    form_class = EntidadReclamoForm
    template_name = 'reclamo/entidadreclamo_form_prueba1.html'


    @ method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate1, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-registro/list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate1,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id
        self.object.estado_reclamo = 0
        return super().form_valid(form)

    def get_context_data(self, **kwargs):

        title = "Actualizar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate1, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco_prueba=paso_cinco_prueba,
            paso_seis_prueba=paso_seis_prueba, paso_siete_prueba=paso_siete_prueba, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)


class EntidadReclamoUpdate_soporte_cerrados(UpdateView):
    model = EntidadReclamo
    form_class = EntidadReclamoForm_secretaria
    template_name = 'reclamo/entidadreclamo_form_ver_detalle_soporte_cerrados.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate_soporte_cerrados, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate_soporte_cerrados,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate_soporte_cerrados, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco=paso_cinco,
            paso_seis=paso_seis, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)

class EntidadReclamoUpdate1(UpdateView):
    success_url = reverse_lazy('reclamo:entidad-reclamo-registro/list')
    model = EntidadReclamo
    form_class = EntidadReclamoForm
    template_name = 'reclamo/entidadreclamo_form_prueba1.html'


    @ method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate1, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-registro/list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate1,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id
        self.object.estado_reclamo = 0
        return super().form_valid(form)

    def get_context_data(self, **kwargs):

        title = "Actualizar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate1, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, paso_cinco_prueba=paso_cinco_prueba,
            paso_seis_prueba=paso_seis_prueba, paso_siete_prueba=paso_siete_prueba, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)



class EntidadReclamoUpdate2(UpdateView):
    success_url = reverse_lazy('reclamo:entidad-reclamo-registro/list')
    model = EntidadReclamo
    form_class = EntidadReclamoForm
    template_name = 'reclamo/entidadreclamo_form_prueba2.html'

    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @ method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate2, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-registro/list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate2,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Actualizar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate2, self).get_context_data(**kwargs), title=title, paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)


class EntidadReclamoUpdate3(UpdateView):
    success_url = reverse_lazy('reclamo:entidad-reclamo-registro/list')
    model = EntidadReclamo
    form_class = EntidadReclamoForm
    template_name = 'reclamo/entidadreclamo_form_prueba3.html'
    # success_url = reverse_lazy('reclamo:entidad-reclamo-list')

    @ method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoUpdate3, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse('reclamo:entidad-reclamo-registro/list', kwargs={}) + "?anio=" + \
            str(self.request.session['reclamo_anio']) + \
            "&mes=" + str(self.request.session['reclamo_mes'])

    def get_form_kwargs(self, *args, **kwargs):
        kwargs = super(EntidadReclamoUpdate3,
                       self).get_form_kwargs(*args, **kwargs)
        entidad_id = self.request.session['entidad_id']
        kwargs['entidad_id'] = entidad_id
        return kwargs

    def form_valid(self, form):
        msg = "Reclamo actulizado correctamente"
        messages.add_message(self.request, messages.SUCCESS, msg)

        self.object = form.save()
        entidad_id = self.request.session['entidad_id']
        self.object.entidad_id = entidad_id
        

        if(self.object.estado_reclamo == 4 ):
            self.object.estado_reclamo = 4
            self.object.etapa_reclamo = 4

        elif(self.object.estado_reclamo ==5 ):
            self.object.estado_reclamo = 5
            self.object.etapa_reclamo = 4

        elif (self.object.estado_reclamo is None and 
        self.object.resultado_reclamo in (3,4,5) and
        self.object.fecha_notificacion is  None
        ):
        
        # Asignar 'etapa_reclamo' igual a 3 y 'estado_reclamo' igual a 1
            self.object.etapa_reclamo = 3
            self.object.estado_reclamo = 1

        elif (self.object.estado_reclamo is None and 
        self.object.resultado_reclamo in (3,4,5) and
        self.object.fecha_notificacion is not None
        ):
        
        # Asignar 'etapa_reclamo' igual a 3 y 'estado_reclamo' igual a 6
            self.object.etapa_reclamo = 4
            self.object.estado_reclamo = 6

        elif (self.object.estado_reclamo is None and 
        self.object.resultado_reclamo in (1,2) and
        self.object.fecha_notificacion is  None
        ):
        
        # Asignar 'etapa_reclamo' igual a 3 y 'estado_reclamo' igual a 1
            self.object.etapa_reclamo = 3
            self.object.estado_reclamo = 1

        elif (self.object.estado_reclamo is None and 
        self.object.resultado_reclamo in (1,2) and 
        self.object.fecha_notificacion is not None
        ):
        
        # Asignar 'etapa_reclamo' igual a 3 y 'estado_reclamo' igual a 1
            self.object.etapa_reclamo = 4
            self.object.estado_reclamo = 6

            # Verificar si 'estado_reclamo' es igual a 1
        elif (self.object.estado_reclamo is None or 
        self.object.resultado_reclamo is None or
        self.object.motivo_conclusion_anticipada is None or
        self.object.fecha_resultado_reclamo is None or
        self.object.comunicacion_resultado_reclamo is None or
        self.object.fecha_notificacion is None):
        
        # Asignar 'etapa_reclamo' y 'estado_reclamo' igual a 2
            self.object.etapa_reclamo = 2
            self.object.estado_reclamo = 2
            self.object.resultado_reclamo = 0



        # Guardar los cambios realizados en el objeto
        self.object.save()
        

        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        title = "Gestionar Reclamo"

        clasificaciones = ClasificacionCausa.objects.all()

        return dict(
            super(EntidadReclamoUpdate3, self).get_context_data(**kwargs), title=title,  paso_uno=paso_uno,
            paso_dos=paso_dos, paso_tres=paso_tres, paso_cuatro=paso_cuatro,  paso_cinco_prueba=paso_cinco_prueba,
            paso_seis_prueba=paso_seis_prueba, paso_siete_prueba=paso_siete_prueba, inputs_hidden=inputs_hidden, clasificaciones=clasificaciones)


class EntidadReclamoDelete(DeleteView):
    success_url = reverse_lazy('reclamo:secretaria-list-nuevos')
    model = EntidadReclamo

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoDelete, self).dispatch(*args, **kwargs)

    def delete(self, request, *args, **kwargs):
        msg = "Ticekt eliminado correctamente"
        messages.add_message(self.request, messages.ERROR,
                             msg, extra_tags='danger')
        return super(EntidadReclamoDelete, self).delete(request, *args, **kwargs)




class EntidadReclamoDelete_programacion(DeleteView):
    model = Programacion
    success_url = reverse_lazy('reclamo:programacion-list-nuevos')
    template_name = 'reclamo/entidadreclamo_confirm_delete_programacion.html'

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoDelete_programacion, self).dispatch(*args, **kwargs)

    def delete(self, request, *args, **kwargs):
        msg = "Programación eliminado correctamente"
        messages.add_message(self.request, messages.ERROR,
                             msg, extra_tags='danger')
        return super(EntidadReclamoDelete_programacion, self).delete(request, *args, **kwargs)


class EntidadReclamoDelete_programacion_entidad(DeleteView):
    model = Programacion
    success_url = reverse_lazy('reclamo:ventas-x-entidad')
    template_name = 'reclamo/entidadreclamo_confirm_delete_programacion.html'

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoDelete_programacion_entidad, self).dispatch(*args, **kwargs)

    def delete(self, request, *args, **kwargs):
        msg = "Programación eliminado correctamente"
        messages.add_message(self.request, messages.ERROR,
                             msg, extra_tags='danger')
        return super(EntidadReclamoDelete_programacion_entidad, self).delete(request, *args, **kwargs)



class EntidadReclamoDelete_monitoreo(DeleteView):
    model = Monitoreo_internet
    success_url = reverse_lazy('reclamo:monitoreo-list-nuevos')
    template_name = 'reclamo/entidadreclamo_confirm_delete_monitoreo.html'

    @method_decorator(valid_access_view(valid_ipress_entidad_edit, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(EntidadReclamoDelete_monitoreo, self).dispatch(*args, **kwargs)

 
    def delete(self, request, *args, **kwargs):
        msg = "Registro eliminado correctamente"
        messages.add_message(self.request, messages.ERROR,
                             msg, extra_tags='danger')
        return super(EntidadReclamoDelete_monitoreo, self).delete(request, *args, **kwargs)  






@require_http_methods(["POST", ])
def guardar_expediente(request):
    id = request.POST['id']
    expediente = request.FILES['expediente']

    try:
        reclamo = EntidadReclamo.objects.get(pk=id)
        reclamo.expediente = expediente
        reclamo.save()
        msg = "Evidencia agregada correctamente"
        messages.add_message(request, messages.SUCCESS, msg)
    except EntidadReclamo.DoesNotExist:
        msg = "Error, Reclamo no existe"
        messages.add_message(request, messages.ERROR, msg, extra_tags='danger')

    return redirect(reverse('reclamo:soporte-list-encurso') + "?anio=" +
                    str(request.session['reclamo_anio']) + "&mes=" + str(request.session['reclamo_mes']))


def guardar_expediente2(request):
    id = request.POST['id']
    expediente2 = request.FILES['expediente2']

    try:
        reclamo = EntidadReclamo.objects.get(pk=id)
        reclamo.expediente2 = expediente2
        reclamo.save()
        msg = "Documento agregado correctamente"
        messages.add_message(request, messages.SUCCESS, msg)
    except EntidadReclamo.DoesNotExist:
        msg = "Error, Reclamo no existe"
        messages.add_message(request, messages.ERROR, msg, extra_tags='danger')

    return redirect(reverse('reclamo:soporte-list-encurso') + "?anio=" +
                    str(request.session['reclamo_anio']) + "&mes=" + str(request.session['reclamo_mes']))

def guardar_expediente_programacion(request):
    id = request.POST['id']
    evidencia = request.FILES['expediente']   

    try:
        reclamo = Programacion.objects.get(pk=id)
        reclamo.evidencia = evidencia
        reclamo.save()
        msg = "Evidencia agregada correctamente"
        messages.add_message(request, messages.SUCCESS, msg)
    except Programacion.DoesNotExist:   
        msg = "Error, programacion no existe"
        messages.add_message(request, messages.ERROR, msg, extra_tags='danger')

    return redirect(
        reverse('reclamo:programacion-list-nuevos')
         
    )



def guardar_expediente_programacion2(request):
    id = request.POST['id']
    documento = request.FILES['expediente2']   

    try:
        reclamo = Programacion.objects.get(pk=id)
        reclamo.documento = documento
        reclamo.save()
        msg = "Documento agregado correctamente"
        messages.add_message(request, messages.SUCCESS, msg)
    except Programacion.DoesNotExist:    
        msg = "Error, programacion no existe"
        messages.add_message(request, messages.ERROR, msg, extra_tags='danger')

    return redirect(
        reverse('reclamo:programacion-list-nuevos')
         
    )



def guardar_expediente_programacion_soporte(request):
    id = request.POST['id']
    evidencia = request.FILES['expediente']   

    try:
        reclamo = Programacion.objects.get(pk=id)
        reclamo.evidencia = evidencia
        reclamo.save()
        msg = "Evidencia agregada correctamente"
        messages.add_message(request, messages.SUCCESS, msg)
    except Programacion.DoesNotExist:   
        msg = "Error, programacion no existe"
        messages.add_message(request, messages.ERROR, msg, extra_tags='danger')

    return redirect(
        reverse('reclamo:programacion-list-nuevos-soporte')
         
    )


def guardar_expediente_programacion2_soporte(request):
    id = request.POST['id']
    documento = request.FILES['expediente2']   

    try:
        reclamo = Programacion.objects.get(pk=id)
        reclamo.documento = documento
        reclamo.save()
        msg = "Documento agregado correctamente"
        messages.add_message(request, messages.SUCCESS, msg)
    except Programacion.DoesNotExist:    
        msg = "Error, programacion no existe"
        messages.add_message(request, messages.ERROR, msg, extra_tags='danger')

    return redirect(
        reverse('reclamo:programacion-list-nuevos-soporte')
         
    )





def atender_reclamo(request, pk):
    reclamo = get_object_or_404(EntidadReclamo, pk=pk)

    if reclamo.id_user:
        messages.error(request, "⚠️Este TICKET ya está en CURSO.")
        return redirect(
            reverse('reclamo:soporte-list-encurso')
            + "?anio=" + str(request.session.get('reclamo_anio', ''))
            + "&mes=" + str(request.session.get('reclamo_mes', ''))
        )

    reclamo.id_user = request.user.id
    reclamo.nombre_soporte = f"{request.user.first_name} {request.user.last_name}"
    reclamo.estado_reclamo = 1
    reclamo.save()

    # ---------------- ENVIO DE CORREO ----------------
    correo_usuario = reclamo.correo_usuario

    if correo_usuario:
        try:
            asunto = "Actualización de su Ticket de Soporte"

            mensaje = f"""
Estimado(a),

Le informamos que su ticket ha sido tomado por el equipo de soporte y se encuentra actualmente en proceso de atención.

Código de Ticket: {reclamo.codigo_ticket}

Personal de soporte asignado:
{reclamo.nombre_soporte}

Nuestro equipo estará trabajando para brindarle una solución a la brevedad posible.

Saludos cordiales,

Oficina de Gestión de Tecnologías de la Información
DIRIS Lima Centro
"""

            send_mail(
                subject=asunto,
                message=mensaje,
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[correo_usuario],
                fail_silently=False,
            )

        except Exception as e:
            messages.warning(
                request,
                "El ticket fue asignado, pero no se pudo enviar el correo de notificación."
            )
    # -------------------------------------------------

    messages.success(request, "Se agregó TICKET en CURSO correctamente.")
    return redirect(
        reverse('reclamo:soporte-list-encurso')
        + "?anio=" + str(request.session.get('reclamo_anio', ''))
        + "&mes=" + str(request.session.get('reclamo_mes', ''))
    )


def dar_por_atendido(request):
    id = request.POST['id']
    
    try:
        reclamo = EntidadReclamo.objects.get(pk=id)
        reclamo.estado_reclamo = 2
        reclamo.fecha_atencion = timezone.now()
        reclamo.save()

        # Enviar correo al usuario
        correo_usuario = reclamo.correo_usuario  # Asegúrate de que este campo exista

        if correo_usuario:
            try:
                asunto = "Su ticket ha sido atendido"
                mensaje = f"""Estimado(a),

Le informamos que su TICKET ha sido atendido correctamente.

Código de Ticket: {reclamo.codigo_ticket}

Gracias por comunicarse con nosotros.

Saludos cordiales,
 
Oficina de Gestión de Tecnologías de la Información - DIRIS LIMA CENTRO
 """
                send_mail(
                    subject=asunto,
                    message=mensaje,
                    from_email=settings.EMAIL_HOST_USER,
                    recipient_list=[correo_usuario],
                    fail_silently=False,
                )
            except Exception as e:
                messages.warning(request, f"El ticket fue atendido, pero no se pudo enviar el correo al usuario.")

        msg = "Ticket ATENDIDO"
        messages.add_message(request, messages.SUCCESS, msg)

    except EntidadReclamo.DoesNotExist:
        msg = "Error, no se atendió el TICKET"
        messages.add_message(request, messages.ERROR, msg, extra_tags='danger')

    return redirect(
        reverse('reclamo:soporte-list-encurso') +
        "?anio=" + str(request.session['reclamo_anio']) +
        "&mes=" + str(request.session['reclamo_mes'])
    )




def valoracion_atencion(request):
    if request.method == 'POST':
        reclamo_id = request.POST.get('id')
        valoracion = request.POST.get('valoracion')  # será "0", "1", o "3"
        observaciones = request.POST.get('observaciones')

        reclamo = get_object_or_404(EntidadReclamo, pk=reclamo_id)

        reclamo.valoracion_atencion = int(valoracion) if valoracion else None
        reclamo.observaciones_atencion = observaciones
        reclamo.estado_reclamo = 3

        reclamo.save()

        messages.success(request, "Se mandó tu valoración MUCHAS GRACIAS 😊")
        return redirect('reclamo:list-cerrados')  # ajusta la redirección a lo que necesites
    

def liberar_caso(request):
    id = request.POST['id']
 
    try:
        reclamo = EntidadReclamo.objects.get(pk=id)
        reclamo.estado_reclamo = 0
        reclamo.id_user = None
        reclamo.nombre_soporte = None
         
        reclamo.save()
        msg = "Ticket Liberado"
        messages.add_message(request, messages.SUCCESS, msg)
    except EntidadReclamo.DoesNotExist:
        msg = "Error, no se atendio el TICKET"
        messages.add_message(request, messages.ERROR, msg, extra_tags='danger')

    return redirect(reverse('reclamo:soporte-list-encurso') + "?anio=" +
                    str(request.session['reclamo_anio']) + "&mes=" + str(request.session['reclamo_mes']))


def guardar_comentario(request):
    if request.method == "POST":
        reclamo_id = request.POST.get('id')
        comentario = request.POST.get("detalle_atencion")

        entidad = get_object_or_404(EntidadReclamo, pk=reclamo_id)
        entidad.detalle_atencion = comentario
        entidad.save()

        messages.success(request, "Comentario guardado correctamente ✅")
        return redirect(request.META.get("HTTP_REFERER", "home"))
    



def dar_por_atendido_programacion(request):
    id = request.POST['id']
 
    try:
        reclamo = Programacion.objects.get(pk=id)
        reclamo.estado_programacion = 1
        reclamo.fecha_atencion = timezone.now() # ← Fecha y hora actual
        reclamo.save()
        msg = "PROGRAMACIÓN ATENDIDA"
        messages.add_message(request, messages.SUCCESS, msg)
    except Programacion.DoesNotExist:
        msg = "Error, no se atendio la programación"
        messages.add_message(request, messages.ERROR, msg, extra_tags='danger')

    return redirect(reverse('reclamo:programacion-list-nuevos') )


def guardar_trabajo(request):
    if request.method == 'POST':
        obj = get_object_or_404(Programacion, pk=request.POST.get('id'))
        obj.trabajo_realizado = request.POST.get('trabajo_realizado')
        obj.recomendaciones = request.POST.get('recomendaciones')
        obj.observaciones = request.POST.get('observaciones')
        obj.save()
        return redirect('reclamo:programacion-list-nuevos')



def guardar_trabajo_programacion(request):
    if request.method == 'POST':
        obj = get_object_or_404(Programacion, pk=request.POST.get('id'))
        obj.trabajo_realizado = request.POST.get('trabajo_realizado')
        obj.recomendaciones = request.POST.get('recomendaciones')
        obj.observaciones = request.POST.get('observaciones')
        obj.save()
        return redirect('reclamo:programacion-list-nuevos-soporte')


def exportar_programacion_pdf(request, id):
    programacion = Programacion.objects.get(pk=id)
    template_path = 'reclamo/acta_programacion.html'

    context = {'p': programacion}

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=acta_{programacion.codigo_programacion}.pdf'

    template = get_template(template_path)
    html = template.render(context)

    # Crear el PDF
    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Error al generar el PDF', status=500)

    return response 



def lista_programaciones(request):
    programaciones = list(Programacion.objects.all())  # conviértelo en lista
    programacion_ids = [p.id for p in programaciones]

    medidas = MedidaAdoptada.objects.filter(entidad_reclamo_id__in=programacion_ids)
    medidas_map = defaultdict(list)
    for m in medidas:
        medidas_map[m.entidad_reclamo_id].append(m)

    # adjuntar la lista de medidas a cada programacion para usar en template
    for p in programaciones:
        p.medidas_vinculadas = medidas_map.get(p.id, [])

    return render(request, "reclamo/entidad_list_programacion.html", {"programaciones": programaciones})



def reporte_programaciones(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    datos = []

    if fecha_inicio and fecha_fin:
        with connection.cursor() as cursor:
            cursor.callproc('obtener_programaciones', [fecha_inicio, fecha_fin])
            resultados = cursor.fetchall()
            columnas = [col[0] for col in cursor.description] if cursor.description else []
        datos = [dict(zip(columnas, fila)) for fila in resultados] if resultados else []

    # Renderizamos el HTML como plantilla
    template = get_template("reclamo/reporte_programaciones.html")
    html = template.render({
        "datos": datos,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
    })

    # Generar el PDF
    response = HttpResponse(content_type="application/pdf")
    response['Content-Disposition'] = 'attachment; filename="PROGRAMACION_REDES_Y_COMUNICACIONES.pdf"'

    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode("UTF-8")), dest=response, encoding='UTF-8')

    if pisa_status.err:
        return HttpResponse("Error al generar el PDF", status=500)

    return response



def reporte_monitoreo(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    tipo_incidencia_param = request.GET.get('tipo_incidencia_param')  # 👈 NUEVO

    datos = []
    total_tiempo = 0
    total_tiempo_texto = ""
    fecha_inicio_formateada = ""
    fecha_fin_formateada = ""

    # Normalizar parámetro
    # "" → None
    # "NULL" → None
    # "1,2,3" → int
    if not tipo_incidencia_param or tipo_incidencia_param == "NULL":
        tipo_incidencia_param = None
    else:
        try:
            tipo_incidencia_param = int(tipo_incidencia_param)
        except:
            tipo_incidencia_param = None

    # configuracion de localización
    try:
        locale.setlocale(locale.LC_TIME, "es_ES.utf8")
    except:
        try:
            locale.setlocale(locale.LC_TIME, "es_ES")
        except:
            locale.setlocale(locale.LC_TIME, "")

    if fecha_inicio and fecha_fin:
        with connection.cursor() as cursor:
            # 👇 LLAMADA AL STORE PROCEDURE CON 3 PARÁMETROS
            cursor.callproc(
                'obtener_monitoreo_por_fecha_de_perdida',
                [fecha_inicio, fecha_fin, tipo_incidencia_param]
            )

            columnas = [col[0] for col in cursor.description]
            datos = [dict(zip(columnas, fila)) for fila in cursor.fetchall()]

        # sumar la columna 8 (índice 7)
        for fila in datos:
            valores = list(fila.values())
            if len(valores) >= 8 and valores[7] is not None:
                try:
                    total_tiempo += float(valores[7])
                except (ValueError, TypeError):
                    pass

        # convertir horas totales a días y horas
        dias = int(total_tiempo // 24)
        horas_restantes = int(total_tiempo % 24)

        if dias == 0:
            total_tiempo_texto = (
                "(1 hora)" if horas_restantes == 1 else f"({horas_restantes} horas)"
            )
        elif horas_restantes == 0:
            total_tiempo_texto = (
                "(1 día)" if dias == 1 else f"({dias} días)"
            )
        else:
            dia_txt = "día" if dias == 1 else "días"
            hora_txt = "hora" if horas_restantes == 1 else "horas"
            total_tiempo_texto = f"({dias} {dia_txt} y {horas_restantes} {hora_txt})"

        # Formateo bonito de fechas
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fecha_fin_dt = datetime.strptime(fecha_fin, "%Y-%m-%d")

            meses = [
                "enero", "febrero", "marzo", "abril", "mayo", "junio",
                "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
            ]
            fecha_inicio_formateada = f"{fecha_inicio_dt.day} de {meses[fecha_inicio_dt.month - 1]} del {fecha_inicio_dt.year}"
            fecha_fin_formateada = f"{fecha_fin_dt.day} de {meses[fecha_fin_dt.month - 1]} del {fecha_fin_dt.year}"
        except:
            fecha_inicio_formateada = fecha_inicio
            fecha_fin_formateada = fecha_fin

    total_tiempo = round(total_tiempo, 2)

    # Renderizar HTML
    template = get_template("reclamo/reporte_monitoreo.html")
    html = template.render({
        "datos": datos,
        "fecha_inicio": fecha_inicio_formateada,
        "fecha_fin": fecha_fin_formateada,
        "total_tiempo": total_tiempo,
        "total_tiempo_texto": total_tiempo_texto,
    })

    # Generar PDF
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="reporte_monitoreo_{fecha_inicio}_a_{fecha_fin}.pdf"'

    pisa_status = pisa.CreatePDF(io.BytesIO(html.encode("UTF-8")), dest=response, encoding="UTF-8")

    if pisa_status.err:
        return HttpResponse("Error al generar el PDF", status=500)

    return response


def obtener_personal_programacion(prog_id):
    with connection.cursor() as cursor:
        cursor.callproc("obtener_personal_programacion", [prog_id])
        result = cursor.fetchall()
    # el SP devuelve una sola fila/columna → lo saco de ahí
    if result:
        return result[0][0]
    return ""


def reporte_programaciones_atendidas_excel(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # Validar fechas
    if not fecha_inicio or not fecha_fin:
        return HttpResponse("Debe proporcionar un rango de fechas válido.", status=400)

    try:
        fecha_inicio = parse_date(fecha_inicio)
        fecha_fin = parse_date(fecha_fin)
    except ValueError:
        return HttpResponse("Formato de fecha inválido.", status=400)

    # Llamar al procedimiento almacenado
    with connection.cursor() as cursor:
        cursor.callproc('obtener_programaciones_atendidas', [fecha_inicio, fecha_fin])
        columnas = [col[0] for col in cursor.description]
        resultados = cursor.fetchall()

    # Crear archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Programaciones Atendidas"

    # Escribir encabezados
    ws.append(columnas)

    # Escribir datos
    for fila in resultados:
        ws.append(fila)

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    nombre_archivo = f"PROGRAMACIONES_ATENDIDAS_{fecha_inicio}_{fecha_fin}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)
    return response


def reporte_monitoreo_internet_excel(request):
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
 
    tipo_incidencia_param = request.GET.get('tipo_incidencia_param')  # NUEVO
 
    # Validar fechas
    if not fecha_inicio or not fecha_fin:
        return HttpResponse("Debe proporcionar un rango de fechas válido.", status=400)

    try:
        fecha_inicio = parse_date(fecha_inicio)
        fecha_fin = parse_date(fecha_fin)
    except ValueError:
        return HttpResponse("Formato de fecha inválido.", status=400)

    # Normalizar parámetro tal como en la otra vista
    if not tipo_incidencia_param or tipo_incidencia_param == "NULL":
        tipo_incidencia_param = None
    else:
        try:
            tipo_incidencia_param = int(tipo_incidencia_param)
        except:
            tipo_incidencia_param = None

    # Llamar al procedimiento almacenado
    with connection.cursor() as cursor:
        cursor.callproc(
            'obtener_monitoreo_por_fecha_de_perdida',
 
            [fecha_inicio, fecha_fin, tipo_incidencia_param]  #  SE AGREGA EL 3ER PARÁMETRO
         )
        columnas = [col[0] for col in cursor.description]
        resultados = cursor.fetchall()

    # Crear archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monitoreo de servicio de internet"

    # Estilos para encabezado
    fill_amarillo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    font_negrita = Font(bold=True)

    # Escribir encabezados
    for idx, col_name in enumerate(columnas, start=1):
        celda = ws.cell(row=1, column=idx, value=col_name)
        celda.fill = fill_amarillo
        celda.font = font_negrita
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max(len(col_name) + 5, 20)

    # Escribir datos
    for fila in resultados:
        ws.append(fila)

    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    nombre_archivo = f"MONITOREO_INTERNET_{fecha_inicio}_{fecha_fin}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

    wb.save(response)
    return response

 
############################################

   
 
class ReporteTicketPDFView(View):
    template_name = "reclamo/reporte_tickets_pdf.html"

    def post(self, request):
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
        soporte = request.user.id

        if not fecha_inicio or not fecha_fin:
            return HttpResponse("Error: Debes seleccionar ambas fechas.")

        # Ejecutar SP
        with connection.cursor() as cursor:
            cursor.callproc('obtener_reporte_tickets', [
                fecha_inicio,
                fecha_fin,
                soporte,
            ])
            desc = cursor.description or []
            filas = cursor.fetchall()

        # Nombres de columnas (sanitizar None -> col_X)
        columnas = []
        for i, col in enumerate(desc):
            name = col[0] if col and col[0] is not None else f"col_{i+1}"
            columnas.append(str(name))

        # Si queremos quitar la columna 11 (índice 10), calculamos el índice real
        idx_eliminar = 10 if len(columnas) > 10 else None
        if idx_eliminar is not None:
            # elimina el header
            columnas.pop(idx_eliminar)

        # Construir filas como lista de listas (todas las celdas string, None -> "")
        rows = []
        for fila in filas:
            fila = list(fila)
            # eliminar columna 11 del resultado bruto si existe
            if idx_eliminar is not None and idx_eliminar < len(fila):
                fila.pop(idx_eliminar)
            # convertir None -> "" y todo a str
            fila_sanitizada = [("" if v is None else str(v)) for v in fila]
            rows.append(fila_sanitizada)

        # DEBUG: escribe la cantidad de columnas/filas al logger (útil para depurar)
        logger.debug("ReporteTicketPDFView -> columnas=%s, filas_count=%s", len(columnas), len(rows))

        # Si no hay columnas o no hay filas, devolver mensaje claro
        if not columnas:
            return HttpResponse("No hay columnas en el resultado del stored procedure.")
        if not rows:
            return HttpResponse("No se encontraron registros para el rango de fechas seleccionado.")

        # Renderizar template con headers y rows
        template = get_template(self.template_name)
        html = template.render({
            "headers": columnas,
            "rows": rows,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "usuario": request.user.get_full_name() or request.user.username,
        })

        # Generar PDF
        response = HttpResponse(content_type="application/pdf")
        response['Content-Disposition'] = 'attachment; filename="reporte_tickets.pdf"'

        pisa_status = pisa.CreatePDF(html, dest=response)

        if pisa_status.err:
            # si falla, lo logueamos y devolvemos mensaje de error
            logger.error("xhtml2pdf error: %s", pisa_status.err)
            return HttpResponse("Hubo un error al generar el PDF.")

        return response




class ReporteTicketPDFView_admin(View):
    template_name = "reclamo/reporte_tickets_pdf_admin.html"

    def post(self, request):
        fecha_inicio = request.POST.get('fecha_inicio')
        fecha_fin = request.POST.get('fecha_fin')
         

        if not fecha_inicio or not fecha_fin:
            return HttpResponse("Error: Debes seleccionar ambas fechas.")

        # Ejecutar SP
        with connection.cursor() as cursor:
            cursor.callproc('obtener_reporte_tickets_admin', [
                fecha_inicio,
                fecha_fin
            ])
            desc = cursor.description or []
            filas = cursor.fetchall()

        # Nombres de columnas (sanitizar None -> col_X)
        columnas = []
        for i, col in enumerate(desc):
            name = col[0] if col and col[0] is not None else f"col_{i+1}"
            columnas.append(str(name))

        # Si queremos quitar la columna 11 (índice 10), calculamos el índice real
        idx_eliminar = 10 if len(columnas) > 10 else None
        if idx_eliminar is not None:
            # elimina el header
            columnas.pop(idx_eliminar)

        # Construir filas como lista de listas (todas las celdas string, None -> "")
        rows = []
        for fila in filas:
            fila = list(fila)
            # eliminar columna 11 del resultado bruto si existe
            if idx_eliminar is not None and idx_eliminar < len(fila):
                fila.pop(idx_eliminar)
            # convertir None -> "" y todo a str
            fila_sanitizada = [("" if v is None else str(v)) for v in fila]
            rows.append(fila_sanitizada)

        # DEBUG: escribe la cantidad de columnas/filas al logger (útil para depurar)
        logger.debug("ReporteTicketPDFView -> columnas=%s, filas_count=%s", len(columnas), len(rows))

        # Si no hay columnas o no hay filas, devolver mensaje claro
        if not columnas:
            return HttpResponse("No hay columnas en el resultado del stored procedure.")
        if not rows:
            return HttpResponse("No se encontraron registros para el rango de fechas seleccionado.")

        # Renderizar template con headers y rows
        template = get_template(self.template_name)
        html = template.render({
            "headers": columnas,
            "rows": rows,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "usuario": request.user.get_full_name() or request.user.username,
        })

        # Generar PDF
        response = HttpResponse(content_type="application/pdf")
        response['Content-Disposition'] = 'attachment; filename="reporte_tickets_admin.pdf"'

        pisa_status = pisa.CreatePDF(html, dest=response)

        if pisa_status.err:
            # si falla, lo logueamos y devolvemos mensaje de error
            logger.error("xhtml2pdf error: %s", pisa_status.err)
            return HttpResponse("Hubo un error al generar el PDF.")

        return response
    

def reporte_tickets_excel(request):
    if request.method == "POST":
        fecha_inicio = request.POST.get("fecha_inicio")
        fecha_fin = request.POST.get("fecha_fin")

        # Crear libro Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Tickets"

        # Ejecutar stored procedure
        with connection.cursor() as cursor:
            cursor.callproc("obtener_reporte_tickets_nuevos", [fecha_inicio, fecha_fin])
            results = cursor.fetchall()

        # ======== ENCABEZADOS PERSONALIZADOS ========
        custom_headers = [
            "Fecha de creación",
            "Estado",
            "Codigo_ticket",
            "Usuario",
            "Celular",
            "Dependencia",
            "Sede",
            "Tipo incidencia",
            "Detalle de solicitud"
        ]

        # ======== ESTILO DE LA CABECERA ========
        header_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")  # Morado
        header_font = Font(bold=True, color="FFFFFF")  # Texto blanco
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Escribir encabezados con estilo
        ws.append(custom_headers)
        for col_num in range(1, len(custom_headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # ======== ESCRIBIR LOS DATOS ========
        for row in results:
            ws.append(row)

        # ======== AJUSTAR ANCHOS DE COLUMNA ========
        for column_cells in ws.columns:
            max_length = 0
            column = column_cells[0].column
            column_letter = get_column_letter(column)

            for cell in column_cells:
                try:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                except:
                    pass

            ws.column_dimensions[column_letter].width = max_length + 5

        # ======== RESPUESTA HTTP ========
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        filename = f"reporte_tickets_nuevos_{fecha_inicio}_a_{fecha_fin}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    return HttpResponse("Método no permitido", status=405)



def atender_reclamo_sihce(request, pk):
    reclamo = get_object_or_404(EntidadReclamo, pk=pk)

    if reclamo.id_user:
        messages.error(request, "⚠️Este TICKET ya está en CURSO.")
        return redirect(
            reverse('reclamo:soporte-list-sihce-encurso')
            + "?anio=" + str(request.session.get('reclamo_anio', ''))
            + "&mes=" + str(request.session.get('reclamo_mes', ''))
        )

    reclamo.id_user = request.user.id
    reclamo.nombre_soporte = request.user.first_name
    reclamo.estado_reclamo = 1
    reclamo.save()

    messages.success(request, "Se agregó TICKET en CURSO correctamente.")
    return redirect(
        reverse('reclamo:soporte-list-sihce-encurso')
        + "?anio=" + str(request.session.get('reclamo_anio', ''))
        + "&mes=" + str(request.session.get('reclamo_mes', ''))
    )

def listar_personas_por_dependencia(request):
    id_dependencia = request.GET.get("id_dependencia")

    if not id_dependencia:
        return JsonResponse({"error": "Falta id_dependencia"}, status=400)

    url = f"http://10.0.5.64/HelpdeskApi/Helpdesk/listarPersona/{id_dependencia}"

    try:
        r = requests.get(url, timeout=5)
        data = r.json()

        resultados = [
            {
                "id": str(p.get("id_persona")),
                "nombre_completo": (
                    f"{p.get('nombre','')} "
                    f"{p.get('apellido_paterno','')} "
                    f"{p.get('apellido_materno','')}"
                ).strip(),
                "cargo": p.get("cargo", ""),
                "telefono": p.get("telefono", ""),   # ✅ NUEVO
                "correo": p.get("correo", "")       # ✅ NUEVO
            }
            for p in data
        ]

        return JsonResponse({"resultados": resultados})

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)





User = get_user_model()

def listar_usuarios(request):
    usuarios = User.objects.filter(rol=3).values(
        "id", "first_name", "last_name"
    )

    data = [
        {
            "id": u["id"],
            "nombre": f"{u['first_name']} {u['last_name']}".strip()
        }
        for u in usuarios
    ]

    return JsonResponse(data, safe=False)


@require_POST
def asignar_reclamo(request):
    reclamo_id = request.POST.get("reclamo_id")
    user_id = request.POST.get("user_id")

    try:
        reclamo = EntidadReclamo.objects.get(id=reclamo_id)
        usuario = User.objects.get(id=user_id)

        # 👇 Como es IntegerField guardamos solo el ID
        reclamo.id_user = usuario.id

        # Guardar nombre del soporte
        reclamo.nombre_soporte = f"{usuario.first_name} {usuario.last_name}"

        reclamo.estado_reclamo = 1
        reclamo.save()

        # ---------------- ENVIO DE CORREO ----------------
        correo_usuario = reclamo.correo_usuario

        if correo_usuario:
            try:
                asunto = "Actualización de su Ticket de Soporte"

                mensaje = f"""
Estimado(a),

Le informamos que su ticket ha sido tomado por el equipo de soporte y se encuentra actualmente en proceso de atención.

Código de Ticket: {reclamo.codigo_ticket}

Personal de soporte asignado:
{reclamo.nombre_soporte}

Nuestro equipo estará trabajando para brindarle una solución a la brevedad posible.

Saludos cordiales,

Oficina de Gestión de Tecnologías de la Información
DIRIS Lima Centro
"""

                send_mail(
                    subject=asunto,
                    message=mensaje,
                    from_email=settings.EMAIL_HOST_USER,
                    recipient_list=[correo_usuario],
                    fail_silently=False,
                )

            except Exception as e:
                print("Error enviando correo:", e)
        # -------------------------------------------------

        return JsonResponse({"success": True})

    except EntidadReclamo.DoesNotExist:
        return JsonResponse({"success": False, "error": "Ticekt no existe"})

    except User.DoesNotExist:
        return JsonResponse({"success": False, "error": "Usuario no existe"})
    


class Mantenimiento_clasificador_ingreso(CreateView):
    form_class = ClasificadoresIngresoForm
    model = Clasificadores_ingreso
    template_name = 'reclamo/mantenimiento_clasificador_ingreso.html'

    @method_decorator(valid_access_view(valid_ipress_entidad_add, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(Mantenimiento_clasificador_ingreso, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        anio = self.request.session.get('reclamo_anio')
        mes = self.request.session.get('reclamo_mes')

        if anio and mes:
            return reverse_lazy('reclamo:clasificador-ingreso') + f"?anio={anio}&mes={mes}"

        return reverse_lazy('reclamo:clasificador-ingreso')

    def form_valid(self, form):
        messages.success(self.request, "Clasificador creado correctamente")
        return super().form_valid(form)

    def get_initial(self):
        initial = super(Mantenimiento_clasificador_ingreso, self).get_initial()

        if self.request.method == 'GET':
            initial.update({
                'periodo': '000000',
                'tipo_institucion': 1,
                'tipo_documento_usuario': 1,
                'tipo_documento_presenta': 1,
                'competencia_reclamo': 1,
                'autorizacion_notificacion_correo': 1
            })

        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['title'] = 'Nuevo Clasificador'

        context['mantenimiento_clasificador_ingreso'] = mantenimiento_clasificador_ingreso

        # 🔥 ESTA LÍNEA HACE QUE SE MUESTREN LOS DATOS EN LA TABLA
        context['clasificadores'] = Clasificadores_ingreso.objects.all().order_by('-id')

        return context


class Mantenimiento_servicios(CreateView):
    form_class = SetupServiciosForm
    model = SetupServicios
    template_name = 'reclamo/mantenimiento_servicios.html'

    @method_decorator(valid_access_view(valid_ipress_entidad_add, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(Mantenimiento_servicios, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        return reverse_lazy('reclamo:servicios')

    def form_valid(self, form):
        messages.success(self.request, "Servicio creado correctamente")
        return super().form_valid(form)

    def get_initial(self):
        initial = super(Mantenimiento_servicios, self).get_initial()

        if self.request.method == 'GET':
            initial.update({
                'descripcion_servicio': '',
                'precio': 0.0,
            })

        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['mantenimiento_servicios'] = mantenimiento_servicios

        # 🔥 ESTA ES LA CLAVE PARA MOSTRAR LA TABLA
        context['servicios'] = SetupServicios.objects.all().order_by('-id')

        context['title'] = 'Nuevo Servicio'

        return context


class Mantenimiento_cuentas_contabilidad(CreateView):
    form_class = ContabilidadCuentaForm
    model = ContabilidadCuenta
    
    template_name = 'reclamo/mantenimiento_cuentas_contabilidad.html'

    @method_decorator(valid_access_view(valid_ipress_entidad_add, login_url='/validate'))
    def dispatch(self, *args, **kwargs):
        return super(Mantenimiento_cuentas_contabilidad, self).dispatch(*args, **kwargs)

    def get_success_url(self):
        anio = self.request.session.get('reclamo_anio')
        mes = self.request.session.get('reclamo_mes')

        if anio and mes:
            return reverse_lazy('reclamo:cuentas-contabilidad') + f"?anio={anio}&mes={mes}"

        return reverse_lazy('reclamo:cuentas-contabilidad')

    def form_valid(self, form):
        messages.success(self.request, "Servicio creado correctamente")
        return super().form_valid(form)

    def get_initial(self):
        initial = super(Mantenimiento_cuentas_contabilidad, self).get_initial()
        if self.request.method == 'GET':
            initial.update({
                'descripcion_servicio': '',
                'precio': 0.0,
            })
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context['mantenimiento_cuentas_contabilidad'] = mantenimiento_cuentas_contabilidad

        context['cuentas'] = ContabilidadCuenta.objects.all().order_by('-id')

        return context



    
    



class Reporte_consolidado_x_clasificador_admin(View):

    def get(self, request):
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        # 🔥 VALIDACIÓN DE FECHAS
        if fecha_inicio and fecha_fin:
            if fecha_inicio > fecha_fin:
                messages.warning(
                    request,
                    "La fecha de inicio no puede ser mayor que la fecha de fin."
                )
                return redirect(request.META.get('HTTP_REFERER', '/'))

        with connection.cursor() as cursor:
            cursor.callproc('sp_reporte_recaudacion', [fecha_inicio, fecha_fin])
            columnas = [col[0] for col in cursor.description]
            datos = [dict(zip(columnas, row)) for row in cursor.fetchall()]

        template = get_template('reclamo/reporte_pdf_consolidado_x_clasificador.html')
        html = template.render({
            'datos': datos,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin
        })

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_recaudacion.pdf"'

        pisa.CreatePDF(html, dest=response)

        return response
    


class Reporte_consolidado_x_clasificador_excel_admin(View):

    def get(self, request):
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        if fecha_inicio and fecha_fin:
            if fecha_inicio > fecha_fin:
                messages.warning(request, "La fecha de inicio no puede ser mayor que la fecha de fin.")
                return redirect(request.META.get('HTTP_REFERER', '/'))

        with connection.cursor() as cursor:
            cursor.callproc('sp_reporte_recaudacion', [fecha_inicio, fecha_fin])
            columnas = [col[0] for col in cursor.description]
            datos = cursor.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # 📦 BORDES
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 🟡 ENCABEZADO AMARILLO
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 📏 ENCABEZADOS
        for col_num, col_name in enumerate(columnas, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name.upper())
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill
            cell.border = thin_border

            # 📏 ancho +50%
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = len(col_name) * 1.5

        # 📊 DATOS
        for row_num, row in enumerate(datos, 2):
            for col_num, value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border

                # alineación
                if col_num == 1:
                    cell.alignment = Alignment(horizontal="left")
                else:
                    cell.alignment = Alignment(horizontal="right")

                # 🔥 última columna en negrita
                if col_num == len(columnas):
                    cell.font = Font(bold=True)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="reporte_recaudacion.xlsx"'

        wb.save(response)
        return response


class Reporte_consolidado_x_clasificador_admin_sismed(View):

    def get(self, request):
        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        # 🔥 VALIDACIÓN DE FECHAS
        if fecha_inicio and fecha_fin:
            if fecha_inicio > fecha_fin:
                messages.warning(
                    request,
                    "La fecha de inicio no puede ser mayor que la fecha de fin."
                )
                return redirect(request.META.get('HTTP_REFERER', '/'))

        with connection.cursor() as cursor:
            cursor.callproc('sp_reporte_recaudacion_sismed', [fecha_inicio, fecha_fin])
            columnas = [col[0] for col in cursor.description]
            datos = [dict(zip(columnas, row)) for row in cursor.fetchall()]

        template = get_template('reclamo/reporte_pdf_consolidado_x_clasificador_sismed.html')
        html = template.render({
            'datos': datos,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin
        })

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="reporte_recaudacion_sismed.pdf"'

        pisa.CreatePDF(html, dest=response)

        return response


 

class Reporte_consolidado_x_clasificador_admin_sismed_excel(View):

    def get(self, request):

        fecha_inicio = request.GET.get(
            'fecha_inicio'
        )

        fecha_fin = request.GET.get(
            'fecha_fin'
        )

        # ======================
        # VALIDAR FECHAS
        # ======================
        if fecha_inicio and fecha_fin:

            if fecha_inicio > fecha_fin:

                messages.warning(
                    request,
                    "La fecha de inicio no puede ser mayor que la fecha de fin."
                )

                return redirect(
                    request.META.get(
                        'HTTP_REFERER',
                        '/'
                    )
                )

        # ======================
        # STORE PROCEDURE
        # ======================
        with connection.cursor() as cursor:

            cursor.callproc(
                'sp_reporte_recaudacion_sismed',
                [
                    fecha_inicio,
                    fecha_fin
                ]
            )

            columnas = [
                col[0]
                for col in cursor.description
            ]

            datos = [
                dict(
                    zip(
                        columnas,
                        row
                    )
                )
                for row in cursor.fetchall()
            ]

        # ======================
        # EXCEL
        # ======================
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "Reporte SISMED"

        # ======================
        # TITULO
        # ======================
        ws.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(columnas)
        )

        ws["A1"] = (
            "REPORTE CONSOLIDADO "
            "POR CLASIFICADOR - SISMED"
        )

        ws["A1"].font = Font(
            bold=True,
            size=14
        )

        ws["A1"].alignment = Alignment(
            horizontal="center"
        )

        # ======================
        # SUBTITULO
        # ======================
        ws.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=len(columnas)
        )

        ws["A2"] = (
            f"Desde: {fecha_inicio}   "
            f"Hasta: {fecha_fin}"
        )

        ws["A2"].alignment = Alignment(
            horizontal="center"
        )

        # ======================
        # ENCABEZADOS
        # ======================
        fila_inicio = 4

        for i, col in enumerate(
            columnas,
            start=1
        ):

            celda = ws.cell(
                row=fila_inicio,
                column=i
            )

            celda.value = col.upper()

            celda.font = Font(
                bold=True,
                color="FFFFFF"
            )

            celda.fill = PatternFill(
                fill_type="solid",
                fgColor="1F4E78"
            )

            celda.alignment = Alignment(
                horizontal="center"
            )

        # ======================
        # DATOS
        # ======================
        fila = fila_inicio + 1

        for row in datos:

            for i, value in enumerate(
                row.values(),
                start=1
            ):

                ws.cell(
                    row=fila,
                    column=i
                ).value = value

            fila += 1

        # ======================
        # ÚLTIMA FILA
        # ======================
        if datos:

            ultima = fila - 1

            for i in range(
                1,
                len(columnas) + 1
            ):

                celda = ws.cell(
                    row=ultima,
                    column=i
                )

                celda.font = Font(
                    bold=True
                )

                celda.fill = PatternFill(
                    fill_type="solid",
                    fgColor="D9D9D9"
                )

        # ======================
        # AUTO ANCHO
        # ======================
        for i in range(
            1,
            len(columnas) + 1
        ):

            letra = get_column_letter(i)

            max_length = 0

            for row in range(
                4,
                ws.max_row + 1
            ):

                valor = ws.cell(
                    row=row,
                    column=i
                ).value

                if valor:
                    largo = len(
                        str(valor)
                    )

                    if largo > max_length:
                        max_length = largo

            ws.column_dimensions[
                letra
            ].width = max_length + 3

        # ======================
        # RESPONSE
        # ======================
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )

        response[
            'Content-Disposition'
        ] = (
            'attachment; '
            'filename="reporte_recaudacion_sismed.xlsx"'
        )

        wb.save(response)

        return response




class CargarDepositoView(View):

    def post(self, request, pk, *args, **kwargs):
        registro = get_object_or_404(
            ConsolidadoDiario,
            pk=pk
        )

        registro.codigo_deposito = request.POST.get(
            'codigo_deposito'
        )

        registro.fecha_deposito = request.POST.get(
            'fecha_deposito'
        )

        if request.FILES.get('constancia_deposito'):
            registro.constancia_deposito = request.FILES.get(
                'constancia_deposito'
            )

        registro.estado = 1

        registro.save()

        messages.success(
            request,
            'Depósito registrado correctamente.'
        )

        return redirect('reclamo:consolidado-diario-admin')
    


class EliminarCuentaContabilidad(DeleteView):
    model = ContabilidadCuenta

    def get_success_url(self):
        messages.success(self.request, "Registro eliminado correctamente")
        return reverse_lazy('reclamo:cuentas-contabilidad')
    

class EliminarServicio(View):

    def post(self, request, pk):
        SetupServicios.objects.filter(pk=pk).delete()

        messages.success(request, "Servicio eliminado correctamente")

        return redirect('reclamo:servicios')
    

class EliminarClasificadorIngreso(View):

    def post(self, request, pk):
        Clasificadores_ingreso.objects.filter(pk=pk).delete()

        messages.success(request, "Clasificador eliminado correctamente")

        return redirect('reclamo:clasificador-ingreso')